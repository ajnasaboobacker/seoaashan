#!/usr/bin/env python3
"""
Google SEO Report Generator - Professional PDF/HTML reports from API data.

Consumes JSON output from seo-google scripts and generates formatted reports
with charts, analytics, and actionable recommendations.

Usage:
    python google_report.py --type cwv-audit --data cwv-data.json --domain example.com
    python google_report.py --type gsc-performance --data gsc-data.json --domain example.com
    python google_report.py --type indexation --data inspect-data.json --domain example.com
    python google_report.py --type full --data full-data.json --domain example.com
    cat data.json | python google_report.py --type cwv-audit --domain example.com
"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    import numpy as np
except ImportError:
    print("Error: matplotlib required. Install with: pip install matplotlib", file=sys.stderr)
    sys.exit(1)

# Prevent WeasyPrint import warnings on Windows from polluting stdout JSON
_stdout_bak = sys.stdout
sys.stdout = sys.stderr
try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except (ImportError, OSError):
    WEASYPRINT_AVAILABLE = False
finally:
    sys.stdout = _stdout_bak


# ─── Brand Colors ────────────────────────────────────────────────────────────

BRAND = {
    "primary": "#1e3a5f",     # Navy (headers, borders)
    "secondary": "#4a5568",    # Warm gray
    "accent": "#b8860b",       # Dark gold
    "success": "#2d6a4f",      # Forest green (pass/good)
    "warning": "#d4740e",      # Warm amber (warnings)
    "danger": "#c53030",       # Deep red (fail/critical)
    "dark": "#1a1a2e",         # Dark navy
    "light_bg": "#faf9f7",     # Warm cream
    "grid": "#d6d3cc",         # Warm border
    "muted": "#6b7280",        # Muted text
}


def _score_color(score):
    """Return brand color based on Lighthouse-style score thresholds."""
    if score >= 90:
        return BRAND["success"]
    elif score >= 50:
        return BRAND["warning"]
    return BRAND["danger"]


def _rating_color(rating):
    """Return brand color based on CrUX rating strings."""
    r = str(rating).lower().replace("-", "_").replace(" ", "_")
    if r in ("good", "pass", "fast"):
        return BRAND["success"]
    elif r in ("needs_improvement", "needs-improvement", "average", "warn"):
        return BRAND["warning"]
    return BRAND["danger"]


def _score_class(score):
    """Return CSS class for TOC score badges."""
    if score >= 80:
        return "score-good"
    elif score >= 50:
        return "score-warn"
    return "score-bad"


def _rating_css_class(rating):
    """Return CSS status class from a rating string."""
    r = str(rating).lower()
    if "good" in r or "pass" in r:
        return "status-pass"
    elif "poor" in r or "fail" in r:
        return "status-fail"
    return "status-warn"


# ─── Chart Setup ─────────────────────────────────────────────────────────────

def _setup_matplotlib():
    """Configure matplotlib for professional chart output."""
    plt.rcParams.update({
        "font.family": "serif",
        "font.serif": ["DejaVu Serif", "Times New Roman", "Georgia"],
        "font.size": 11,
        "axes.titlesize": 14,
        "axes.titleweight": "bold",
        "axes.labelsize": 11,
        "axes.facecolor": "white",
        "figure.facecolor": "white",
        "axes.grid": False,
        "axes.spines.top": False,
        "axes.spines.right": False,
    })


_setup_matplotlib()


# ─── Chart Functions ─────────────────────────────────────────────────────────

def chart_lighthouse_gauges(data: dict, output_dir: Path) -> str:
    """Generate 2x2 Lighthouse score gauges."""
    scores = data.get("lighthouse_scores", {})
    if not scores:
        return ""

    fig, axes = plt.subplots(2, 2, figsize=(8, 4), subplot_kw={"projection": "polar"})
    categories = [
        ("performance", "Performance"),
        ("accessibility", "Accessibility"),
        ("best-practices", "Best Practices"),
        ("seo", "SEO"),
    ]

    for ax, (key, label) in zip(axes.flat, categories):
        score = scores.get(key, 0)
        theta_bg = np.linspace(np.pi, 0, 100)
        theta_fill = np.linspace(np.pi, np.pi - (score / 100) * np.pi, 100)

        ax.plot(theta_bg, [1] * 100, linewidth=14, color="#e2e8f0", solid_capstyle="round")
        ax.plot(theta_fill, [1] * 100, linewidth=14, color=_score_color(score), solid_capstyle="round")

        ax.text(np.pi / 2, 0.35, f"{score}", ha="center", va="center",
                fontsize=24, fontweight="bold", color=BRAND["dark"])
        ax.text(np.pi / 2, -0.05, label, ha="center", va="center",
                fontsize=10, color=BRAND["muted"])

        ax.set_ylim(0, 1.3)
        ax.set_rticks([])
        ax.set_thetagrids([])
        ax.spines["polar"].set_visible(False)

    plt.tight_layout(pad=2)
    path = output_dir / "lighthouse_gauges.png"
    plt.savefig(path, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close()
    return str(path)


def chart_cwv_distributions(data: dict, output_dir: Path) -> str:
    """Generate stacked horizontal bars for CWV metric distributions."""
    crux = data.get("crux", {})
    metrics = crux.get("metrics", {})
    if not metrics:
        return ""

    cwv_order = [
        "largest_contentful_paint", "interaction_to_next_paint",
        "cumulative_layout_shift", "first_contentful_paint",
        "experimental_time_to_first_byte",
    ]

    labels, goods, nis, poors = [], [], [], []
    for name in cwv_order:
        m = metrics.get(name)
        if not m or "distribution" not in m:
            continue
        d = m["distribution"]
        labels.append(m.get("label", name))
        goods.append(d.get("good", 0))
        nis.append(d.get("needs_improvement", 0))
        poors.append(d.get("poor", 0))

    if not labels:
        return ""

    fig, ax = plt.subplots(figsize=(8, max(2.5, len(labels) * 0.7)))
    y = range(len(labels))

    ax.barh(y, goods, color=BRAND["success"], label="Good", height=0.5)
    ax.barh(y, nis, left=goods, color=BRAND["warning"], label="Needs Improvement", height=0.5)
    left2 = [g + n for g, n in zip(goods, nis)]
    ax.barh(y, poors, left=left2, color=BRAND["danger"], label="Poor", height=0.5)

    ax.set_yticks(y)
    ax.set_yticklabels(labels)
    ax.set_xlim(0, 100)
    ax.set_xlabel("% of page loads")
    ax.legend(loc="lower right", fontsize=9)
    ax.invert_yaxis()

    for i, (g, n, p) in enumerate(zip(goods, nis, poors)):
        if g > 10:
            ax.text(g / 2, i, f"{g:.0f}%", ha="center", va="center",
                    fontsize=8, color="white", fontweight="bold")
        if n > 10:
            ax.text(g + n / 2, i, f"{n:.0f}%", ha="center", va="center",
                    fontsize=8, color="white", fontweight="bold")
        if p > 10:
            ax.text(g + n + p / 2, i, f"{p:.0f}%", ha="center", va="center",
                    fontsize=8, color="white", fontweight="bold")

    plt.tight_layout()
    path = output_dir / "cwv_distributions.png"
    plt.savefig(path, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close()
    return str(path)


def chart_cwv_timeline(data: dict, output_dir: Path) -> str:
    """Generate CWV timeline chart from CrUX History data."""
    metrics = data.get("metrics", {})
    periods = data.get("collection_periods", [])
    if not metrics or not periods:
        return ""

    cwv_metrics = [
        "largest_contentful_paint",
        "interaction_to_next_paint",
        "cumulative_layout_shift",
    ]
    available = [m for m in cwv_metrics if m in metrics]
    if not available:
        return ""

    fig, axes = plt.subplots(len(available), 1, figsize=(10, 3 * len(available)), sharex=True)
    if len(available) == 1:
        axes = [axes]

    x_labels = [p.get("last", "")[-5:] for p in periods]  # MM-DD format
    x = range(len(x_labels))

    for ax, metric_name in zip(axes, available):
        m = metrics[metric_name]
        p75s = m.get("p75_values", [])
        label = m.get("label", metric_name)
        good_t = m.get("good_threshold", 0)
        poor_t = m.get("poor_threshold", 0)

        valid_x = [i for i, v in enumerate(p75s) if v is not None]
        valid_y = [v for v in p75s if v is not None]

        if not valid_y:
            continue

        # Threshold bands
        if good_t and poor_t:
            ax.axhspan(0, good_t, alpha=0.1, color=BRAND["success"])
            ax.axhspan(good_t, poor_t, alpha=0.1, color=BRAND["warning"])
            ax.axhline(y=good_t, color=BRAND["success"], linestyle="--", alpha=0.5, linewidth=1)
            ax.axhline(y=poor_t, color=BRAND["danger"], linestyle="--", alpha=0.5, linewidth=1)

        ax.plot(valid_x, valid_y, color=BRAND["primary"], linewidth=2, marker="o", markersize=3)
        ax.fill_between(valid_x, valid_y, alpha=0.1, color=BRAND["primary"])

        unit = m.get("unit", "")
        ax.set_ylabel(f"{label} (p75{unit})")
        ax.set_title(label, fontsize=12, fontweight="bold")

    if x_labels:
        step = max(1, len(x_labels) // 8)
        axes[-1].set_xticks(range(0, len(x_labels), step))
        axes[-1].set_xticklabels(
            [x_labels[i] for i in range(0, len(x_labels), step)],
            rotation=45, fontsize=8,
        )

    plt.tight_layout()
    path = output_dir / "cwv_timeline.png"
    plt.savefig(path, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close()
    return str(path)


def chart_top_queries(data: dict, output_dir: Path) -> str:
    """Generate horizontal bar chart of top queries by impressions."""
    rows = data.get("rows", [])
    if not rows:
        return ""

    # Sort by impressions (more meaningful than clicks for new sites)
    top = sorted(rows, key=lambda r: r.get("impressions", 0), reverse=True)[:12]
    top = [r for r in top if r.get("impressions", 0) > 0]
    if not top:
        return ""

    labels = [r.get("query", r.get("keys", ["?"])[0])[:35] for r in top]
    impressions = [r.get("impressions", 0) for r in top]
    clicks = [r.get("clicks", 0) for r in top]

    if not impressions or max(impressions) < 3:
        return ""

    fig, ax = plt.subplots(figsize=(7, max(2, len(labels) * 0.3)))
    y = range(len(labels))
    bars = ax.barh(y, impressions, color=BRAND["primary"], height=0.55, label="Impressions")
    if any(c > 0 for c in clicks):
        ax.barh(y, clicks, color=BRAND["success"], height=0.55, label="Clicks")
        ax.legend(fontsize=8, loc="lower right")
    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=8)
    ax.set_xlabel("Count")
    ax.invert_yaxis()

    for bar, val in zip(bars, clicks):
        if val > 0:
            ax.text(
                bar.get_width() + max(clicks) * 0.02,
                bar.get_y() + bar.get_height() / 2,
                str(val), va="center", fontsize=8, color=BRAND["dark"],
            )

    plt.tight_layout()
    path = output_dir / "top_queries.png"
    plt.savefig(path, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close()
    return str(path)


def chart_index_status(data: dict, output_dir: Path) -> str:
    """Generate donut chart for URL inspection results."""
    summary = data.get("summary", {})
    if not summary:
        return ""

    labels, sizes, colors = [], [], []
    for key, label, color in [
        ("pass", "Indexed", BRAND["success"]),
        ("fail", "Not Indexed", BRAND["danger"]),
        ("neutral", "Neutral", BRAND["grid"]),
        ("error", "Error", BRAND["muted"]),
    ]:
        val = summary.get(key, 0)
        if val > 0:
            labels.append(f"{label} ({val})")
            sizes.append(val)
            colors.append(color)

    if not sizes:
        return ""

    fig, ax = plt.subplots(figsize=(4.5, 3.5))
    wedges, texts, autotexts = ax.pie(
        sizes, labels=labels, colors=colors, autopct="%1.0f%%",
        startangle=90, pctdistance=0.75, textprops={"fontsize": 9},
    )
    centre = plt.Circle((0, 0), 0.50, fc="white")
    ax.add_artist(centre)
    total = sum(sizes)
    ax.text(0, 0, f"{total}\nURLs", ha="center", va="center",
            fontsize=16, fontweight="bold", color=BRAND["dark"])

    plt.tight_layout()
    path = output_dir / "index_status.png"
    plt.savefig(path, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close()
    return str(path)


def chart_crawl_categories(data: dict, output_dir: Path) -> str:
    """Generate horizontal bar chart of category scores for crawl audit."""
    categories = {
        "technical": "Technical SEO",
        "content": "Content Quality",
        "onpage": "On-Page SEO",
        "schema": "Schema Markup",
        "performance": "Performance (CWV)",
        "geo": "AI Search (GEO)",
        "images": "Image Optimization"
    }
    scores_dict = data.get("categoryScores", {})
    if not scores_dict:
        return ""

    labels = []
    scores = []
    colors = []
    for key, label in categories.items():
        if key in scores_dict:
            score = scores_dict[key]
            labels.append(label)
            scores.append(score)
            colors.append(_score_color(score))

    if not scores:
        return ""

    fig, ax = plt.subplots(figsize=(7, 3.5))
    y = range(len(labels))
    bars = ax.barh(y, scores, color=colors, height=0.5)

    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=9, fontweight="bold", color=BRAND["dark"])
    ax.set_xlim(0, 100)
    ax.set_xlabel("Score", fontsize=10, color=BRAND["muted"])
    ax.invert_yaxis()

    # Add values to the end of bars
    for bar, score in zip(bars, scores):
        ax.text(
            bar.get_width() + 2,
            bar.get_y() + bar.get_height() / 2,
            f"{score}/100",
            va="center",
            ha="left",
            fontsize=9,
            fontweight="bold",
            color=BRAND["dark"]
        )

    plt.tight_layout()
    path = output_dir / "crawl_categories.png"
    plt.savefig(path, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close()
    return str(path)


# ─── CSS Template ────────────────────────────────────────────────────────────

def _build_css(domain: str) -> str:
    """
    Professional modern report CSS.
    Uses clean sans-serif typography, gradients, neon accents, and responsive layout classes.
    """
    return f"""\
  @page {{
    size: A4;
    margin: 22mm 18mm 25mm 18mm;
    @bottom-center {{
      content: "Page " counter(page) " of " counter(pages);
      font-size: 8.5pt;
      color: #64748b;
      font-family: system-ui, -apple-system, sans-serif;
    }}
    @bottom-left {{
      content: "Confidential // Tier 4 SEO Audit";
      font-size: 7.5pt;
      color: #94a3b8;
      font-family: system-ui, -apple-system, sans-serif;
    }}
    @bottom-right {{
      content: "{domain} SEO Audit Report";
      font-size: 8pt;
      color: #94a3b8;
      font-family: system-ui, -apple-system, sans-serif;
    }}
  }}

  @page :first {{
    margin: 0;
    @bottom-left {{ content: none; }}
    @bottom-center {{ content: none; }}
    @bottom-right {{ content: none; }}
  }}

  @page toc {{
    @bottom-center {{
      content: counter(page);
      font-size: 9pt;
      color: #64748b;
    }}
  }}

  * {{
    box-sizing: border-box;
    margin: 0;
    padding: 0;
  }}

  body {{
    font-family: system-ui, -apple-system, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
    font-size: 9.5pt;
    line-height: 1.5;
    color: #1e293b;
    background: white;
  }}

  /* ─── Title Page (Premium Cyber Dark Mode) ─── */
  .title-page {{
    page: first;
    width: 210mm;
    height: 297mm;
    background: linear-gradient(135deg, #090c0d 0%, #141a1c 100%);
    display: flex;
    flex-direction: column;
    justify-content: center;
    align-items: center;
    text-align: center;
    color: #e5f9f6;
    position: relative;
    padding: 50mm 20mm 40mm 20mm;
    border-top: 6mm solid #00ff66;
    border-bottom: 6mm solid #00f0ff;
  }}

  .title-page .badge {{
    background: rgba(0, 255, 102, 0.08);
    border: 1px solid rgba(0, 255, 102, 0.2);
    border-radius: 20px;
    padding: 6px 18px;
    font-size: 8.5pt;
    letter-spacing: 2px;
    text-transform: uppercase;
    margin-bottom: 12mm;
    color: #00ff66;
    font-family: monospace;
    font-weight: bold;
    box-shadow: 0 0 10px rgba(0, 255, 102, 0.05);
  }}

  .title-page h1 {{
    font-size: 26pt;
    font-weight: 800;
    margin-bottom: 4mm;
    letter-spacing: -1px;
    line-height: 1.1;
    color: #ffffff;
    font-family: system-ui, -apple-system, sans-serif;
  }}

  .title-page .subtitle {{
    font-size: 12pt;
    color: #b8cfcd;
    margin-bottom: 12mm;
    font-weight: 400;
    letter-spacing: 0.5px;
    text-transform: uppercase;
  }}

  .title-page .url {{
    font-size: 11pt;
    color: #00f0ff;
    margin-bottom: 15mm;
    padding: 3mm 6mm;
    border: 1px solid rgba(0, 240, 255, 0.2);
    border-radius: 4px;
    background: rgba(0, 240, 255, 0.04);
    font-family: monospace;
  }}

  .title-page .score-box {{
    background: rgba(255, 255, 255, 0.01);
    border: 2px solid rgba(0, 255, 102, 0.3);
    border-radius: 12px;
    padding: 6mm 12mm;
    margin-bottom: 15mm;
    box-shadow: 0 0 15px rgba(0, 255, 102, 0.1);
  }}

  .title-page .score-number {{
    font-size: 42pt;
    font-weight: bold;
    color: #00ff66;
    line-height: 1;
    text-shadow: 0 0 10px rgba(0, 255, 102, 0.3);
  }}

  .title-page .score-label {{
    font-size: 9.5pt;
    color: #b8cfcd;
    margin-top: 2mm;
    text-transform: uppercase;
    letter-spacing: 1px;
  }}

  .title-page .meta {{
    font-size: 8.5pt;
    color: #5c7573;
    margin-top: 8mm;
    padding-top: 5mm;
    border-top: 1px solid #1e2629;
    width: 60%;
  }}

  .title-page .meta span {{
    margin: 0 8px;
  }}

  /* ─── Table of Contents ─── */
  .toc-page {{
    page: toc;
    page-break-before: always;
    padding-top: 10mm;
  }}

  .toc-page h2 {{
    font-size: 16pt;
    color: #0f172a;
    margin-bottom: 8mm;
    padding-bottom: 3mm;
    border-bottom: 2px solid #0f172a;
    font-weight: 700;
  }}

  .toc-list {{
    list-style: none;
    padding: 0;
  }}

  .toc-list li {{
    padding: 2.5mm 0;
    border-bottom: 1px solid #f1f5f9;
    overflow: hidden;
  }}

  .toc-list li.toc-section {{
    font-weight: bold;
    font-size: 10.5pt;
    padding-top: 3.5mm;
    color: #0f172a;
  }}

  .toc-list li.toc-sub {{
    padding-left: 6mm;
    font-size: 9pt;
    color: #475569;
  }}

  .toc-score {{
    display: inline-block;
    float: right;
    padding: 2px 8px;
    border-radius: 12px;
    font-size: 8pt;
    font-weight: bold;
    color: white;
  }}

  .score-good {{ background: #16a34a; }}
  .score-warn {{ background: #d97706; }}
  .score-bad {{ background: #dc2626; }}

  /* ─── Section Styles ─── */
  div.section {{
    page-break-before: always;
  }}

  .section-header {{
    background: #f8fafc;
    border-left: 4px solid #0f172a;
    padding: 5mm 6mm;
    margin-bottom: 6mm;
    page-break-after: avoid;
    border-radius: 0 6px 6px 0;
  }}

  .section-header h2 {{
    font-size: 15pt;
    color: #0f172a;
    margin-bottom: 1.5mm;
    font-weight: 700;
  }}

  .section-header .section-score {{
    font-size: 11pt;
    font-weight: bold;
    float: right;
    margin-top: -6mm;
    color: #0f172a;
  }}

  h3 {{
    font-size: 11pt;
    color: #0f172a;
    margin-top: 6mm;
    margin-bottom: 3mm;
    padding-bottom: 1.5mm;
    border-bottom: 1px solid #e2e8f0;
    font-weight: 700;
    page-break-after: avoid;
  }}

  h4 {{
    font-size: 10pt;
    color: #334155;
    margin-top: 4mm;
    margin-bottom: 2mm;
    font-weight: 600;
    page-break-after: avoid;
  }}

  p {{
    margin-bottom: 2.5mm;
    color: #334155;
  }}

  .highlight {{
    background: #fffbeb;
    border-left: 3px solid #ffb700;
    padding: 3mm 4mm;
    margin: 4mm 0;
    font-size: 9pt;
    border-radius: 0 4px 4px 0;
  }}

  .critical-box {{
    background: #fef2f2;
    border-left: 3px solid #ff3c5c;
    padding: 3mm 4mm;
    margin: 4mm 0;
    font-size: 9pt;
    border-radius: 0 4px 4px 0;
  }}

  .success-box {{
    background: #f0fdf4;
    border-left: 3px solid #00ff66;
    padding: 3mm 4mm;
    margin: 4mm 0;
    font-size: 9pt;
    border-radius: 0 4px 4px 0;
  }}

  /* ─── Tables ─── */
  table {{
    width: 100%;
    border-collapse: collapse;
    margin: 4mm 0 6mm 0;
    font-size: 8.5pt;
  }}

  thead th {{
    background: #f8fafc;
    color: #0f172a;
    font-weight: bold;
    padding: 2.5mm 3.5mm;
    text-align: left;
    border-bottom: 2px solid #e2e8f0;
    font-size: 8.5pt;
  }}

  tbody td {{
    padding: 2mm 3mm;
    border-bottom: 1px solid #f1f5f9;
    vertical-align: top;
  }}

  tbody tr:nth-child(even) {{
    background: #fafbfc;
  }}

  .status-pass {{
    color: #16a34a;
    font-weight: bold;
  }}

  .status-fail {{
    color: #dc2626;
    font-weight: bold;
  }}

  .status-warn {{
    color: #d97706;
    font-weight: bold;
  }}

  .status-partial {{
    color: #475569;
    font-weight: bold;
  }}

  /* ─── Charts ─── */
  .chart-container {{
    text-align: center;
    margin: 5mm 0;
  }}

  .chart-container img {{
    max-width: 100%;
    max-height: 120mm;
    height: auto;
  }}

  .chart-caption {{
    font-size: 8pt;
    color: #64748b;
    font-style: italic;
    margin-top: 2mm;
    text-align: center;
  }}

  .chart-half {{
    display: inline-block;
    width: 48%;
    vertical-align: top;
    text-align: center;
    margin: 2mm 0;
  }}

  .chart-half img {{
    max-width: 100%;
    height: auto;
  }}

  /* ─── Two column layout ─── */
  .two-col {{
    display: table;
    width: 100%;
    table-layout: fixed;
    margin: 3mm 0;
  }}

  .two-col .col {{
    display: table-cell;
    vertical-align: top;
    padding: 0 2mm;
  }}

  .four-col {{
    display: table;
    width: 100%;
    table-layout: fixed;
    margin: 3mm 0;
  }}

  .four-col .col {{
    display: table-cell;
    vertical-align: top;
    padding: 0 1.5mm;
  }}

  /* ─── Metric Cards ─── */
  .metric-card {{
    background: #f8fafc;
    border: 1px solid #e2e8f0;
    border-radius: 6px;
    padding: 2.5mm 3mm;
    text-align: center;
    margin: 2mm 0;
    box-shadow: 0 1px 2px rgba(0,0,0,0.01);
  }}

  .metric-card .value {{
    font-size: 13pt;
    font-weight: bold;
    line-height: 1.2;
    color: #0f172a;
  }}

  .metric-card .label {{
    font-size: 7pt;
    color: #64748b;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-top: 1mm;
  }}

  /* ─── Action Plan ─── */
  .action-item {{
    background: #f8fafc;
    border-radius: 6px;
    padding: 3mm 4mm;
    margin: 3mm 0;
    border: 1px solid #e2e8f0;
    border-left: 3px solid #94a3b8;
  }}

  .action-item.critical {{
    border-left-color: #ff3c5c;
    background: #fef2f2;
    border-color: #fee2e2;
  }}

  .action-item.high {{
    border-left-color: #ffb700;
    background: #fffbeb;
    border-color: #fef3c7;
  }}

  .action-item.medium {{
    border-left-color: #00f0ff;
    background: #f0fdfa;
    border-color: #ccfbf1;
  }}

  .action-item.low {{
    border-left-color: #00ff66;
    background: #f0fdf4;
    border-color: #bbf7d0;
  }}

  .action-item h4 {{
    margin-top: 0;
    margin-bottom: 1.5mm;
    border-bottom: none;
    padding-bottom: 0;
    font-weight: bold;
  }}

  .action-item .effort {{
    font-size: 8pt;
    color: #64748b;
    float: right;
  }}

  .priority-tag {{
    display: inline-block;
    padding: 0.5mm 3mm;
    border-radius: 3px;
    font-size: 7.5pt;
    font-weight: bold;
    color: white;
    margin-right: 2mm;
    vertical-align: middle;
  }}

  .priority-critical {{ background: #ff3c5c; }}
  .priority-high {{ background: #ffb700; }}
  .priority-medium {{ background: #00f0ff; color: #0f172a; }}
  .priority-low {{ background: #00ff66; color: #0f172a; }}

  /* ─── Code blocks ─── */
  .code-block {{
    background: #0f172a;
    color: #e2e8f0;
    padding: 3mm 4mm;
    border-radius: 6px;
    font-family: 'DejaVu Sans Mono', monospace;
    font-size: 7.5pt;
    line-height: 1.5;
    margin: 3mm 0;
    white-space: pre-wrap;
    word-break: break-all;
    border: 1px solid #1e293b;
  }}

  /* ─── Divider ─── */
  .divider {{
    border: none;
    border-top: 1px solid #e2e8f0;
    margin: 5mm 0;
  }}

  /* ─── Roadmap ─── */
  .roadmap-phase {{
    background: #f8fafc;
    border-radius: 6px;
    padding: 4mm 5mm;
    margin: 4mm 0;
    border: 1px solid #e2e8f0;
  }}

  .roadmap-phase h4 {{
    margin-top: 0;
    border-bottom: none;
    color: #0f172a;
  }}

  .roadmap-phase ul {{
    margin: 2mm 0 0 5mm;
    padding: 0;
  }}

  .roadmap-phase li {{
    margin-bottom: 1.5mm;
    font-size: 9pt;
    color: #334155;
  }}

  /* ─── Lists ─── */
  ul {{
    margin-left: 5mm;
    margin-bottom: 3mm;
  }}

  li {{
    margin-bottom: 1.5mm;
  }}

  /* ─── Data freshness ─── */
  .data-freshness {{
    font-size: 7.5pt;
    color: #64748b;
    font-style: italic;
    margin-top: 4mm;
    padding-top: 2mm;
    border-top: 1px solid #e2e8f0;
  }}

  @media screen {{
    body {{
      background-color: #090c0d;
      padding: 40px 10px;
      display: flex;
      flex-direction: column;
      align-items: center;
      font-size: 10pt;
    }}
    .title-page, .toc-page, div.section {{
      background-color: #ffffff;
      width: 210mm;
      min-height: 297mm;
      padding: 22mm 18mm 25mm 18mm;
      margin: 0 auto 30px auto;
      border-radius: 8px;
      box-shadow: 0 15px 35px rgba(0, 0, 0, 0.45);
      position: relative;
    }}
    .title-page {{
      display: flex;
      flex-direction: column;
      justify-content: center;
      align-items: center;
      height: 297mm;
      background: linear-gradient(135deg, #090c0d 0%, #141a1c 100%);
      border-top: 6mm solid #00ff66;
      border-bottom: 6mm solid #00f0ff;
      padding-top: 0;
      padding-bottom: 0;
    }}
  }}
"""


# ─── HTML Helpers ────────────────────────────────────────────────────────────

def _img_tag(path, alt="Chart"):
    """Convert file path to file:// URI img tag for WeasyPrint."""
    if not path:
        return ""
    return f'<img src="file://{path}" style="max-width: 100%;" alt="{alt}">'


def _chart_html(path, caption, fig_num, alt="Chart"):
    """Build a complete chart container with figure caption."""
    if not path:
        return ""
    return (
        f'    <div class="chart-container">\n'
        f'      <img src="file://{path}" style="width: 85%;" alt="{alt}">\n'
        f'      <div class="chart-caption">Figure {fig_num}: {caption}</div>\n'
        f'    </div>\n'
    )


def _metric_card(value, label, color=None):
    """Build a metric card HTML block."""
    style = f' style="color: {color};"' if color else ""
    return (
        f'      <div class="metric-card">\n'
        f'        <div class="value"{style}>{value}</div>\n'
        f'        <div class="label">{label}</div>\n'
        f'      </div>\n'
    )


# ─── Section Builders ────────────────────────────────────────────────────────

def _build_title_page(domain, report_title, subtitle, score=None, score_label=None, meta_items=None):
    """Build the gradient title page."""
    score_html = ""
    if score is not None:
        score_html = (
            f'  <div class="score-box">\n'
            f'    <div class="score-number">{score}'
            f'<span style="font-size: 20pt; color: #94a3b8;">/100</span></div>\n'
            f'    <div class="score-label">{score_label or "Lighthouse Performance Score"}</div>\n'
            f'  </div>\n'
        )

    meta_html = ""
    if meta_items:
        spans = "\n    ".join(f"<span>{item}</span>" for item in meta_items)
        meta_html = (
            f'  <div class="meta">\n'
            f'    {spans}\n'
            f'  </div>\n'
        )

    # Google logo (if available in charts dir)
    google_logo_path = Path(__file__).parent.parent / "charts" / "google_logo.png"
    google_logo_html = ""
    if google_logo_path.exists():
        google_logo_html = (
            f'  <div style="margin-top: 8mm;">\n'
            f'    <img src="file://{google_logo_path}" style="height: 20px; opacity: 0.7;" alt="Google">\n'
            f'    <div style="font-size: 7pt; color: #6b7280; margin-top: 2mm;">Powered by Google APIs</div>\n'
            f'  </div>\n'
        )

    return (
        f'\n<!-- {"=" * 55} TITLE PAGE {"=" * 3} -->\n'
        f'<div class="title-page">\n'
        f'  <div class="badge">{report_title}</div>\n'
        f'  <h1>{domain}</h1>\n'
        f'  <div class="subtitle">Prepared by SEO AASHAN</div>\n'
        f'{score_html}'
        f'{meta_html}'
        f'{google_logo_html}'
        f'</div>\n'
    )


def _build_toc(sections_info):
    """
    Build a Table of Contents page.

    sections_info: list of dicts with keys 'num', 'title', 'score' (optional),
                   and 'subs' (list of subtitle strings).
    """
    items = []
    for sec in sections_info:
        score_html = ""
        if sec.get("score") is not None:
            cls = _score_class(sec["score"])
            score_html = f' <span class="toc-score {cls}">{sec["score"]}</span>'
        items.append(
            f'    <li class="toc-section">'
            f'<span>{sec["num"]}. {sec["title"]}</span>{score_html}'
            f'</li>'
        )
        for sub in sec.get("subs", []):
            items.append(f'    <li class="toc-sub"><span>{sub}</span></li>')

    items_html = "\n".join(items)
    return (
        f'\n<!-- {"=" * 55} TABLE OF CONTENTS {"=" * 3} -->\n'
        f'<div class="toc-page">\n'
        f'  <h2>Table of Contents</h2>\n'
        f'  <ul class="toc-list">\n'
        f'{items_html}\n'
        f'  </ul>\n'
        f'</div>\n'
    )


def _build_executive_summary(domain, timestamp, data, report_type):
    """Build the Executive Summary section with metric cards, issues, and wins."""
    lines = []
    lines.append(f'\n<!-- {"=" * 55} 1. EXECUTIVE SUMMARY {"=" * 3} -->')
    lines.append('<div class="section">')
    lines.append('  <div class="section-header">')
    lines.append('    <h2>1. Executive Summary</h2>')
    lines.append('  </div>')
    lines.append('')

    # Context paragraph
    lines.append(f'  <p>This report presents a comprehensive Google SEO analysis of '
                 f'<strong>{domain}</strong>, generated on {timestamp}. '
                 f'Data was collected from Google PageSpeed Insights, Chrome User Experience '
                 f'Report (CrUX), Google Search Console, and the URL Inspection API as available.</p>')
    lines.append('')

    # Metric cards row
    cards = []

    # PSI performance score
    psi = data.get("psi", {})
    mobile = psi.get("psi", {}).get("mobile", psi) if isinstance(psi, dict) else {}
    perf_score = mobile.get("lighthouse_scores", {}).get("performance")
    if perf_score is not None:
        color = _score_color(perf_score)
        cards.append(("perf", f"{perf_score}/100", "Lighthouse Performance", color))

    seo_score = mobile.get("lighthouse_scores", {}).get("seo")
    if seo_score is not None:
        color = _score_color(seo_score)
        cards.append(("seo", f"{seo_score}/100", "Lighthouse SEO", color))

    # GSC totals
    gsc = data.get("gsc", {})
    if gsc.get("totals"):
        clicks = gsc["totals"].get("clicks", 0)
        cards.append(("clicks", f"{clicks:,}", "Total Clicks", BRAND["primary"]))
        impr = gsc["totals"].get("impressions", 0)
        cards.append(("impr", f"{impr:,}", "Impressions", BRAND["secondary"]))

    # Indexation
    inspection = data.get("inspection", {})
    if inspection.get("summary"):
        indexed = inspection["summary"].get("pass", 0)
        total = inspection.get("total", 0)
        cards.append(("idx", f"{indexed}/{total}", "Indexed URLs", BRAND["success"]))

    # Render metric cards in a compact row
    if cards:
        col_class = "four-col" if len(cards) >= 4 else "two-col"
        display_cards = cards[:5]
        lines.append(f'  <div class="{col_class}">')
        for _, val, lbl, clr in display_cards:
            lines.append(f'    <div class="col">')
            lines.append(_metric_card(val, lbl, clr))
            lines.append(f'    </div>')
        lines.append('  </div>')
        lines.append('')

    # Critical issues box
    issues = []
    failed_audits = mobile.get("failed_audits", [])
    if failed_audits:
        top_fail = sorted(failed_audits, key=lambda a: a.get("score", 1))[:3]
        for a in top_fail:
            issues.append(f'<strong>{a.get("title", "Unknown")}</strong> '
                          f'(score: {a.get("score", 0):.0%})')

    seo_audits = mobile.get("seo_audits", [])
    seo_failed = [a for a in seo_audits if not a.get("pass")]
    for a in seo_failed[:2]:
        issues.append(f'<strong>SEO:</strong> {a.get("title", "Unknown")}')

    inspect_fails = inspection.get("summary", {}).get("fail", 0)
    if inspect_fails:
        issues.append(f'<strong>{inspect_fails} URL(s)</strong> not indexed')

    if issues:
        issue_items = "\n".join(f"      <li>{i}</li>" for i in issues[:5])
        lines.append('  <div class="critical-box">')
        lines.append('    <strong>Critical Issues Found:</strong>')
        lines.append('    <ol>')
        lines.append(issue_items)
        lines.append('    </ol>')
        lines.append('  </div>')
        lines.append('')

    # Quick wins box
    wins = []
    qw = gsc.get("quick_wins", [])
    if qw:
        wins.append(f'{len(qw)} search queries at positions 4-10 with high impressions '
                    f'(small ranking bump = significant traffic)')

    opps = mobile.get("opportunities", [])
    for o in opps[:3]:
        savings = o.get("savings_ms", 0)
        if savings:
            wins.append(f'{o.get("title", "Optimization")}: save ~{savings}ms')

    if wins:
        win_items = "\n".join(f"      <li>{w}</li>" for w in wins[:5])
        lines.append('  <div class="success-box">')
        lines.append('    <strong>Quick Wins:</strong>')
        lines.append('    <ol>')
        lines.append(win_items)
        lines.append('    </ol>')
        lines.append('  </div>')
        lines.append('')

    lines.append('</div>')
    return "\n".join(lines)


def _build_cwv_section(psi_data, crux_data, chart_paths, history_data=None, section_num=2):
    """Build the Core Web Vitals audit section."""
    fig_counter = [1]  # mutable counter for figure numbering

    def next_fig():
        n = fig_counter[0]
        fig_counter[0] += 1
        return n

    lines = []
    lines.append(f'\n<!-- {"=" * 55} {section_num}. CORE WEB VITALS {"=" * 3} -->')
    lines.append('<div class="section">')
    lines.append('  <div class="section-header">')
    lines.append(f'    <h2>{section_num}. Core Web Vitals &amp; Performance</h2>')

    # Section score from Lighthouse performance
    psi = psi_data if isinstance(psi_data, dict) else {}
    mobile = psi.get("psi", {}).get("mobile", psi)
    scores = mobile.get("lighthouse_scores", {})
    perf = scores.get("performance")
    if perf is not None:
        color = _score_color(perf)
        lines.append(f'    <div class="section-score" style="color: {color};">{perf}/100</div>')

    lines.append('  </div>')
    lines.append('')

    # 2.1 Lighthouse Scores gauges
    if scores:
        lines.append(f'  <h3>{section_num}.1 Lighthouse Scores</h3>')
        lines.append('')
        gauges_path = chart_paths.get("gauges_path", "")
        lines.append(_chart_html(
            gauges_path,
            "Lighthouse audit scores across Performance, Accessibility, Best Practices, and SEO.",
            next_fig(),
            "Lighthouse gauge scores",
        ))

        # Scores summary table
        lines.append('  <table>')
        lines.append('    <thead>')
        lines.append('      <tr><th>Category</th><th>Score</th><th>Rating</th></tr>')
        lines.append('    </thead>')
        lines.append('    <tbody>')
        for key, label in [("performance", "Performance"), ("accessibility", "Accessibility"),
                           ("best-practices", "Best Practices"), ("seo", "SEO")]:
            s = scores.get(key)
            if s is not None:
                cls = "status-pass" if s >= 90 else ("status-warn" if s >= 50 else "status-fail")
                rating = "Good" if s >= 90 else ("Needs Work" if s >= 50 else "Poor")
                lines.append(f'      <tr><td>{label}</td><td class="{cls}">{s}</td><td>{rating}</td></tr>')
        lines.append('    </tbody>')
        lines.append('  </table>')
        lines.append('')

    # Divider between Lighthouse Scores and Lab Metrics
    lines.append('  <hr class="divider">')

    # 2.2 Lab Metrics
    lab = mobile.get("lab_metrics", {})
    if lab:
        lines.append(f'  <h3>{section_num}.2 Lab Metrics (Simulated)</h3>')
        lines.append('  <table>')
        lines.append('    <thead>')
        lines.append('      <tr><th>Metric</th><th>Value</th><th>Score</th><th>Threshold</th></tr>')
        lines.append('    </thead>')
        lines.append('    <tbody>')
        metric_labels = {
            "first-contentful-paint": "First Contentful Paint (FCP)",
            "largest-contentful-paint": "Largest Contentful Paint (LCP)",
            "total-blocking-time": "Total Blocking Time (TBT)",
            "cumulative-layout-shift": "Cumulative Layout Shift (CLS)",
            "speed-index": "Speed Index",
            "interactive": "Time to Interactive (TTI)",
        }
        thresholds = {
            "first-contentful-paint": "\u2264 1.8s",
            "largest-contentful-paint": "\u2264 2.5s",
            "total-blocking-time": "\u2264 200ms",
            "cumulative-layout-shift": "\u2264 0.1",
            "speed-index": "\u2264 3.4s",
            "interactive": "\u2264 3.8s",
        }
        for k, v in lab.items():
            score_val = v.get("score")
            score_pct = f"{score_val:.0%}" if score_val is not None else "N/A"
            cls = ("status-pass" if score_val and score_val >= 0.9
                   else ("status-warn" if score_val and score_val >= 0.5 else "status-fail"))
            label = metric_labels.get(k, k.replace("-", " ").title())
            threshold = thresholds.get(k, "\u2014")
            lines.append(f'      <tr><td>{label}</td><td>{v.get("display", "")}</td>'
                         f'<td class="{cls}">{score_pct}</td><td>{threshold}</td></tr>')
        lines.append('    </tbody>')
        lines.append('  </table>')
        lines.append('')

    # 2.3 CrUX Field Data
    crux = crux_data if isinstance(crux_data, dict) else {}
    crux_metrics = crux.get("metrics", {})
    if crux_metrics:
        lines.append(f'  <h3>{section_num}.3 CrUX Field Data (28-day Rolling Average)</h3>')
        lines.append('')

        # Distribution chart
        dist_path = chart_paths.get("distributions_path", "")
        lines.append(_chart_html(
            dist_path,
            "Core Web Vitals field data distribution across Good, Needs Improvement, and Poor buckets.",
            next_fig(),
            "CWV distribution chart",
        ))

        lines.append('  <table>')
        lines.append('    <thead>')
        lines.append('      <tr><th>Metric</th><th>p75</th><th>Rating</th>'
                     '<th>Good %</th><th>NI %</th><th>Poor %</th></tr>')
        lines.append('    </thead>')
        lines.append('    <tbody>')
        for name, m in crux_metrics.items():
            rating = m.get("rating", "?")
            dist = m.get("distribution", {})
            unit = m.get("unit", "")
            p75 = m.get("p75", "?")
            display_val = f"{p75:.3f}" if name == "cumulative_layout_shift" else f"{p75}{unit}"
            cls = _rating_css_class(rating)
            lines.append(f'      <tr><td>{m.get("label", name)}</td><td>{display_val}</td>')
            lines.append(f'      <td class="{cls}">{rating.upper()}</td>')
            lines.append(f'      <td>{dist.get("good", "N/A")}%</td>'
                         f'<td>{dist.get("needs_improvement", "N/A")}%</td>'
                         f'<td>{dist.get("poor", "N/A")}%</td></tr>')
        lines.append('    </tbody>')
        lines.append('  </table>')

        cp = crux.get("collection_period", {})
        if cp:
            lines.append(f'  <p class="data-freshness">Collection period: {cp.get("first", "?")} '
                         f'to {cp.get("last", "?")}. CrUX data is a 28-day rolling average '
                         f'updated daily ~04:00 UTC.</p>')
        lines.append('')
    elif crux.get("error"):
        lines.append(f'  <h3>{section_num}.3 CrUX Field Data</h3>')
        lines.append(f'  <div class="highlight"><strong>CrUX Field Data:</strong> '
                     f'{crux["error"]}</div>')
        lines.append('')

    # CrUX History timeline
    if history_data and not history_data.get("error"):
        lines.append(f'  <h3>{section_num}.4 Core Web Vitals Trends (25-week)</h3>')
        timeline_path = chart_paths.get("timeline_path", "")
        lines.append(_chart_html(
            timeline_path,
            "CrUX p75 values over 25 weeks with Good/Poor threshold bands.",
            next_fig(),
            "CWV timeline trends",
        ))

        trends = history_data.get("trends", {})
        if trends:
            lines.append('  <table>')
            lines.append('    <thead>')
            lines.append('      <tr><th>Metric</th><th>Direction</th><th>Change</th>'
                         '<th>Earliest Avg</th><th>Latest Avg</th></tr>')
            lines.append('    </thead>')
            lines.append('    <tbody>')
            for name, t in trends.items():
                direction = t.get("direction", "?")
                cls = ("status-pass" if direction == "improving"
                       else ("status-fail" if direction == "degrading" else ""))
                lines.append(f'      <tr><td>{t.get("label", name)}</td>'
                             f'<td class="{cls}">{direction.upper()}</td>')
                lines.append(f'      <td>{t.get("change_pct", 0):+.1f}%</td>'
                             f'<td>{t.get("earliest_avg", "?")}</td>'
                             f'<td>{t.get("latest_avg", "?")}</td></tr>')
            lines.append('    </tbody>')
            lines.append('  </table>')
        lines.append('')

    # Failed audits
    failed = mobile.get("failed_audits", [])
    if failed:
        sub = f"{section_num}.5" if history_data and not history_data.get("error") else f"{section_num}.4"
        lines.append(f'  <h3>{sub} Failed / Warning Audits ({len(failed)})</h3>')
        lines.append('  <table>')
        lines.append('    <thead>')
        lines.append('      <tr><th>Audit</th><th>Score</th><th>Details</th></tr>')
        lines.append('    </thead>')
        lines.append('    <tbody>')
        for a in failed[:20]:
            score_pct = f"{a['score']:.0%}" if a.get("score") is not None else "?"
            lines.append(f'      <tr><td>{a.get("title", "")}</td>'
                         f'<td class="status-fail">{score_pct}</td>'
                         f'<td>{a.get("display", "")}</td></tr>')
        lines.append('    </tbody>')
        lines.append('  </table>')
        lines.append('')

    # SEO audit checks
    seo_audits = mobile.get("seo_audits", [])
    if seo_audits:
        seo_failed = [a for a in seo_audits if not a.get("pass")]
        if seo_failed:
            lines.append(f'  <h3>SEO Audit Issues ({len(seo_failed)})</h3>')
            for a in seo_failed:
                lines.append(f'  <div class="action-item critical">')
                lines.append(f'    <h4>{a.get("title", "")}</h4>')
                lines.append(f'  </div>')
        else:
            lines.append(f'  <div class="success-box"><strong>SEO:</strong> '
                         f'All {len(seo_audits)} Lighthouse SEO checks passed.</div>')
        lines.append('')

    # Accessibility issues
    a11y = mobile.get("accessibility_audits", [])
    if a11y:
        lines.append(f'  <h3>Accessibility Issues ({len(a11y)})</h3>')
        lines.append('  <table>')
        lines.append('    <thead>')
        lines.append('      <tr><th>Issue</th><th>Score</th></tr>')
        lines.append('    </thead>')
        lines.append('    <tbody>')
        for a in a11y:
            lines.append(f'      <tr><td>{a.get("title", "")}</td>'
                         f'<td class="status-fail">{a.get("score", 0):.0%}</td></tr>')
        lines.append('    </tbody>')
        lines.append('  </table>')
        lines.append('')

    # Opportunities
    opps = mobile.get("opportunities", [])
    if opps:
        lines.append(f'  <h3>Optimization Opportunities ({len(opps)})</h3>')
        lines.append('  <table>')
        lines.append('    <thead>')
        lines.append('      <tr><th>Opportunity</th><th>Estimated Savings</th></tr>')
        lines.append('    </thead>')
        lines.append('    <tbody>')
        for o in opps:
            lines.append(f'      <tr><td>{o.get("title", "")}</td>'
                         f'<td>{o.get("savings_ms", 0)}ms</td></tr>')
        lines.append('    </tbody>')
        lines.append('  </table>')
        lines.append('')

    lines.append('</div>')
    return "\n".join(lines), fig_counter[0]


def _build_gsc_section(gsc_data, chart_paths, section_num=3, fig_start=1):
    """Build the GSC Search Performance section."""
    fig_counter = [fig_start]

    def next_fig():
        n = fig_counter[0]
        fig_counter[0] += 1
        return n

    lines = []
    lines.append(f'\n<!-- {"=" * 55} {section_num}. SEARCH PERFORMANCE {"=" * 3} -->')
    lines.append('<div class="section">')
    lines.append('  <div class="section-header">')
    lines.append(f'    <h2>{section_num}. Search Console Performance</h2>')
    lines.append('  </div>')
    lines.append('')

    totals = gsc_data.get("totals", {})
    dr = gsc_data.get("date_range", {})

    if totals:
        domain = gsc_data.get("property", "?")
        lines.append(f'  <p>Period: {dr.get("start", "?")} to {dr.get("end", "?")} '
                     f'| Property: {domain}</p>')
        queries_count = gsc_data.get("row_count", 0)
        impr_total = totals.get("impressions", 0)
        lines.append(f'  <p><strong>{domain}</strong> appeared in <strong>{queries_count}</strong> unique search queries '
                     f'with <strong>{impr_total:,}</strong> total impressions during this period.</p>')
        lines.append('')

        clicks_val = f'{totals.get("clicks", 0):,}'
        impr_val = f'{totals.get("impressions", 0):,}'
        ctr_val = f'{totals.get("ctr", 0)}%'
        rows_val = str(gsc_data.get("row_count", 0))

        # Metric cards in single row
        lines.append(f'  <h3>{section_num}.1 Key Metrics</h3>')
        lines.append('  <div class="four-col">')
        lines.append(f'    <div class="col">{_metric_card(clicks_val, "Total Clicks", BRAND["primary"])}</div>')
        lines.append(f'    <div class="col">{_metric_card(impr_val, "Total Impressions", BRAND["secondary"])}</div>')
        lines.append(f'    <div class="col">{_metric_card(ctr_val, "Average CTR", BRAND["accent"])}</div>')
        lines.append(f'    <div class="col">{_metric_card(rows_val, "Queries Found", BRAND["dark"])}</div>')
        lines.append('  </div>')
        lines.append('  <hr class="divider">')
        lines.append('')

    # Top queries chart
    queries_path = chart_paths.get("top_queries_path", "")
    if queries_path:
        lines.append(f'  <h3>{section_num}.2 Top Queries by Impressions</h3>')
        lines.append(_chart_html(
            queries_path,
            "Top search queries ranked by impression volume from Google Search Console (28-day period).",
            next_fig(),
            "Top queries bar chart",
        ))

    # Top queries table
    rows = gsc_data.get("rows", [])
    if rows:
        lines.append(f'  <h3>{section_num}.3 Query Detail Table</h3>')
        lines.append('  <table>')
        lines.append('    <thead>')
        lines.append('      <tr><th>#</th><th>Query</th><th>Clicks</th>'
                     '<th>Impressions</th><th>CTR</th><th>Position</th></tr>')
        lines.append('    </thead>')
        lines.append('    <tbody>')
        sorted_rows = sorted(rows, key=lambda r: r.get("impressions", 0), reverse=True)
        for i, r in enumerate(sorted_rows[:15], 1):
            query = r.get("query", r.get("keys", ["?"])[0])
            pos = r.get("position", 0)
            pos_cls = ("status-pass" if pos <= 3
                       else ("status-warn" if pos <= 10 else "status-fail"))
            lines.append(f'      <tr><td>{i}</td><td>{query}</td>'
                         f'<td>{r.get("clicks", 0)}</td>'
                         f'<td>{r.get("impressions", 0):,}</td>')
            lines.append(f'      <td>{r.get("ctr", 0)}%</td>'
                         f'<td class="{pos_cls}">{pos:.1f}</td></tr>')
        lines.append('    </tbody>')
        lines.append('  </table>')
        lines.append('')

    # Position analysis
    if rows:
        top3 = len([r for r in rows if r.get("position", 99) <= 3])
        top10 = len([r for r in rows if r.get("position", 99) <= 10])
        beyond = len([r for r in rows if r.get("position", 99) > 10])
        lines.append(f'  <h3>{section_num}.4 Query Position Analysis</h3>')
        lines.append('  <div class="two-col">')
        lines.append(f'    <div class="col">')
        lines.append(_metric_card(str(top3), "Queries in Top 3", BRAND["success"]))
        lines.append(f'    </div>')
        lines.append(f'    <div class="col">')
        lines.append(_metric_card(str(top10), "Queries in Top 10", BRAND["warning"]))
        lines.append(f'    </div>')
        lines.append('  </div>')
        if beyond:
            lines.append(f'  <p>{beyond} queries rank beyond position 10 '
                         f'and may benefit from content optimization.</p>')
        lines.append('')

    # Quick wins
    qw = gsc_data.get("quick_wins", [])
    if qw:
        lines.append(f'  <h3>{section_num}.5 Quick Wins ({len(qw)} opportunities)</h3>')
        lines.append('  <div class="highlight">These queries rank at position 4-10 with '
                     'high impressions. A small ranking improvement could yield significant '
                     'traffic gains.</div>')
        lines.append('  <table>')
        lines.append('    <thead>')
        lines.append('      <tr><th>Query</th><th>Position</th>'
                     '<th>Impressions</th><th>Clicks</th></tr>')
        lines.append('    </thead>')
        lines.append('    <tbody>')
        for w in qw:
            query = w.get("keys", ["?"])[0] if w.get("keys") else w.get("query", "?")
            lines.append(f'      <tr><td>{query}</td>'
                         f'<td>{w.get("position", 0):.1f}</td>'
                         f'<td>{w.get("impressions", 0):,}</td>'
                         f'<td>{w.get("clicks", 0)}</td></tr>')
        lines.append('    </tbody>')
        lines.append('  </table>')
        lines.append('')

    lines.append('  <p class="data-freshness">Search Analytics data has a 2-3 day lag. '
                 'Data available for ~16 months.</p>')
    lines.append('</div>')
    return "\n".join(lines), fig_counter[0]


def _build_indexation_section(inspect_data, chart_paths, section_num=4, fig_start=1):
    """Build the Indexation Status section."""
    fig_counter = [fig_start]

    def next_fig():
        n = fig_counter[0]
        fig_counter[0] += 1
        return n

    lines = []
    lines.append(f'\n<!-- {"=" * 55} {section_num}. INDEXATION STATUS {"=" * 3} -->')
    lines.append('<div class="section">')
    lines.append('  <div class="section-header">')
    lines.append(f'    <h2>{section_num}. Indexation Status</h2>')
    lines.append('  </div>')
    lines.append('')

    summary = inspect_data.get("summary", {})
    total = inspect_data.get("total", 0)

    if summary:
        # Donut chart
        idx_path = chart_paths.get("index_status_path", "")
        if idx_path:
            fig_n = next_fig()
            lines.append(f'  <h3>{section_num}.1 Index Coverage Overview</h3>')
            lines.append(f'    <div class="chart-container">')
            lines.append(f'      <img src="file://{idx_path}" style="width: 70%;" alt="Index status donut chart">')
            lines.append(f'      <div class="chart-caption">Figure {fig_n}: URL indexation status distribution from Google URL Inspection API.</div>')
            lines.append(f'    </div>')

        # Summary cards
        lines.append(f'  <p>Total URLs inspected: <strong>{total}</strong></p>')
        lines.append('  <div class="two-col">')
        lines.append(f'    <div class="col">')
        lines.append(_metric_card(summary.get("pass", 0), "Indexed", BRAND["success"]))
        lines.append(f'    </div>')
        lines.append(f'    <div class="col">')
        lines.append(_metric_card(summary.get("fail", 0), "Not Indexed", BRAND["danger"]))
        lines.append(f'    </div>')
        lines.append('  </div>')
        lines.append('')

        indexed = summary.get("pass", 0)
        not_indexed = summary.get("fail", 0)
        if total > 0:
            rate = round((indexed / total) * 100, 1)
            lines.append(f'  <p><strong>Index Rate:</strong> {rate}% of inspected URLs are indexed by Google.</p>')
        if total > 0 and indexed > 0:
            pct = (indexed / total) * 100
            cls = "success-box" if pct >= 90 else ("highlight" if pct >= 70 else "critical-box")
            lines.append(f'  <div class="{cls}"><strong>Index Rate:</strong> '
                         f'{pct:.0f}% of inspected URLs are indexed.</div>')
            lines.append('')

    # Per-URL results table
    results = inspect_data.get("results", [])
    if results:
        lines.append(f'  <h3>{section_num}.2 Per-URL Results</h3>')
        lines.append('  <table>')
        lines.append('    <thead>')
        lines.append('      <tr><th>URL</th><th>Verdict</th>'
                     '<th>Coverage State</th><th>Last Crawl</th></tr>')
        lines.append('    </thead>')
        lines.append('    <tbody>')
        for r in results:
            verdict = r.get("verdict", "?")
            cls = ("status-pass" if verdict == "PASS"
                   else ("status-fail" if verdict == "FAIL" else ""))
            idx = r.get("index_status", {})
            cov = idx.get("coverage_state", r.get("error", "N/A"))
            crawl = idx.get("last_crawl_time", "N/A")
            if crawl and crawl != "N/A":
                crawl = crawl[:10]
            url_display = r.get("url", "?")
            lines.append(f'      <tr><td style="word-break:break-all;font-size:8pt;">'
                         f'{url_display}</td>')
            lines.append(f'      <td class="{cls}">{verdict}</td>'
                         f'<td>{cov}</td><td>{crawl}</td></tr>')
        lines.append('    </tbody>')
        lines.append('  </table>')
        lines.append('')

    # Rich results
    rich = inspect_data.get("rich_results", [])
    if rich:
        lines.append(f'  <h3>{section_num}.3 Rich Results Detected</h3>')
        lines.append('  <table>')
        lines.append('    <thead>')
        lines.append('      <tr><th>URL</th><th>Rich Result Type</th></tr>')
        lines.append('    </thead>')
        lines.append('    <tbody>')
        for rr in rich:
            lines.append(f'      <tr><td style="word-break:break-all;font-size:8pt;">'
                         f'{rr.get("url", "?")}</td>'
                         f'<td>{rr.get("type", "?")}</td></tr>')
        lines.append('    </tbody>')
        lines.append('  </table>')
        lines.append('')

    lines.append('  <p class="data-freshness">URL Inspection API: '
                 '2,000 inspections/day per property.</p>')
    lines.append('</div>')
    return "\n".join(lines), fig_counter[0]


def _build_recommendations(data, section_num=5):
    """Build prioritized recommendations section based on discovered issues."""
    lines = []
    lines.append(f'\n<!-- {"=" * 55} {section_num}. RECOMMENDATIONS {"=" * 3} -->')
    lines.append('<div class="section">')
    lines.append('  <div class="section-header">')
    lines.append(f'    <h2>{section_num}. Recommendations</h2>')
    lines.append('  </div>')
    lines.append('')
    lines.append('  <p>Prioritized action items based on the data collected. '
                 'Items are ranked by expected impact on search visibility and user experience.</p>')
    lines.append('')

    item_num = 0

    # Collect critical items
    critical_items = []
    psi = data.get("psi", {})
    mobile = psi.get("psi", {}).get("mobile", psi) if isinstance(psi, dict) else {}

    perf = mobile.get("lighthouse_scores", {}).get("performance")
    if perf is not None and perf < 50:
        critical_items.append(
            ("Improve Lighthouse Performance Score", "Medium (2-4 hrs)",
             f"Current score is {perf}/100. Focus on reducing Largest Contentful Paint "
             f"and Total Blocking Time. Defer non-critical JavaScript and optimize images.")
        )

    seo_failed = [a for a in mobile.get("seo_audits", []) if not a.get("pass")]
    for a in seo_failed[:2]:
        critical_items.append(
            (f"Fix SEO Issue: {a.get('title', 'Unknown')}", "Low (30 min)",
             "This Lighthouse SEO check is failing. Address it to ensure proper crawling "
             "and indexing by search engines.")
        )

    inspect = data.get("inspection", {})
    not_indexed = inspect.get("summary", {}).get("fail", 0)
    if not_indexed:
        critical_items.append(
            (f"Resolve {not_indexed} Non-Indexed URL(s)", "Medium (2-4 hrs)",
             "These pages are not appearing in Google's index. Review coverage state, "
             "fix crawl errors, and request re-indexing via Search Console.")
        )

    if critical_items:
        lines.append(f'  <h3><span class="priority-tag priority-critical">CRITICAL</span> '
                     f'Fix Immediately</h3>')
        for title, effort, desc in critical_items:
            item_num += 1
            lines.append(f'  <div class="action-item critical">')
            lines.append(f'    <h4>{item_num}. {title} '
                         f'<span class="effort">Effort: {effort}</span></h4>')
            lines.append(f'    <p>{desc}</p>')
            lines.append(f'  </div>')
        lines.append('')

    # High priority items
    high_items = []
    failed_audits = mobile.get("failed_audits", [])
    top_fails = sorted(failed_audits, key=lambda a: a.get("score", 1))[:5]
    for a in top_fails:
        if a.get("score", 1) < 0.5:
            high_items.append(
                (f"Address: {a.get('title', 'Unknown')}", "Medium (1-2 hrs)",
                 f"Score: {a['score']:.0%}. {a.get('display', 'Review and optimize this audit.')}")
            )

    opps = mobile.get("opportunities", [])
    for o in opps[:3]:
        savings = o.get("savings_ms", 0)
        if savings:
            high_items.append(
                (f"{o.get('title', 'Optimization')}", "Medium (2-4 hrs)",
                 f"Potential savings of ~{savings}ms. Implement this to improve page load speed.")
            )

    gsc = data.get("gsc", {})
    qw = gsc.get("quick_wins", [])
    if qw:
        high_items.append(
            (f"Optimize {len(qw)} Quick-Win Queries", "Medium (2-4 hrs)",
             "These queries rank at positions 4-10 with high impressions. "
             "Improve on-page SEO and content depth to push them into top 3.")
        )

    if high_items:
        lines.append(f'  <h3><span class="priority-tag priority-high">HIGH</span> '
                     f'Fix Within 1 Week</h3>')
        for title, effort, desc in high_items:
            item_num += 1
            lines.append(f'  <div class="action-item high">')
            lines.append(f'    <h4>{item_num}. {title} '
                         f'<span class="effort">Effort: {effort}</span></h4>')
            lines.append(f'    <p>{desc}</p>')
            lines.append(f'  </div>')
        lines.append('')

    # Medium priority items
    medium_items = []
    a11y = mobile.get("accessibility_audits", [])
    if a11y:
        medium_items.append(
            (f"Fix {len(a11y)} Accessibility Issue(s)", "Low (30 min)",
             "Accessibility improvements benefit SEO (Lighthouse score) and user experience. "
             "Address failing accessibility audits.")
        )

    acc_score = mobile.get("lighthouse_scores", {}).get("accessibility")
    if acc_score is not None and acc_score < 90 and not a11y:
        medium_items.append(
            ("Improve Accessibility Score", "Medium (2-4 hrs)",
             f"Current score: {acc_score}/100. Run a detailed accessibility audit "
             f"and address any violations.")
        )

    bp_score = mobile.get("lighthouse_scores", {}).get("best-practices")
    if bp_score is not None and bp_score < 90:
        medium_items.append(
            ("Address Best Practices Issues", "Low (30 min)",
             f"Current score: {bp_score}/100. Review browser console for errors, "
             f"update deprecated APIs, and ensure HTTPS for all resources.")
        )

    if medium_items:
        lines.append(f'  <h3><span class="priority-tag priority-medium">MEDIUM</span> '
                     f'Fix Within 1 Month</h3>')
        for title, effort, desc in medium_items:
            item_num += 1
            lines.append(f'  <div class="action-item medium">')
            lines.append(f'    <h4>{item_num}. {title} '
                         f'<span class="effort">Effort: {effort}</span></h4>')
            lines.append(f'    <p>{desc}</p>')
            lines.append(f'  </div>')
        lines.append('')

    # If no recommendations were generated at all
    if item_num == 0:
        lines.append('  <div class="success-box"><strong>No critical issues detected.</strong> '
                     'Continue monitoring Core Web Vitals and search performance regularly.</div>')
        lines.append('')

    # Implementation Roadmap
    lines.append('  <hr class="divider">')
    lines.append('  <h3>Implementation Roadmap</h3>')
    lines.append('  <div class="roadmap-phase">')
    lines.append('    <h4>Week 1 &mdash; Quick Wins</h4>')
    lines.append('    <ul>')
    if seo_failed:
        lines.append('      <li>Fix failing Lighthouse SEO checks</li>')
    if a11y:
        lines.append(f'      <li>Address {len(a11y)} accessibility issue(s)</li>')
    bp_score_val = mobile.get("lighthouse_scores", {}).get("best-practices")
    if bp_score_val is not None and bp_score_val < 90:
        lines.append('      <li>Review and fix Best Practices issues</li>')
    if not seo_failed and not a11y and (bp_score_val is None or bp_score_val >= 90):
        lines.append('      <li>Verify all monitoring dashboards are active</li>')
    lines.append('    </ul>')
    lines.append('  </div>')
    lines.append('  <div class="roadmap-phase">')
    lines.append('    <h4>Week 2&ndash;3 &mdash; Performance &amp; Indexation</h4>')
    lines.append('    <ul>')
    if perf is not None and perf < 50:
        lines.append('      <li>Optimize Largest Contentful Paint and Total Blocking Time</li>')
    if not_indexed:
        lines.append(f'      <li>Resolve {not_indexed} non-indexed URL(s)</li>')
    if opps:
        lines.append(f'      <li>Implement {len(opps)} performance optimization(s)</li>')
    if (perf is None or perf >= 50) and not not_indexed and not opps:
        lines.append('      <li>Maintain current performance levels and monitor trends</li>')
    lines.append('    </ul>')
    lines.append('  </div>')
    lines.append('  <div class="roadmap-phase">')
    lines.append('    <h4>Week 4 &mdash; Content &amp; Search Optimization</h4>')
    lines.append('    <ul>')
    if qw:
        lines.append(f'      <li>Optimize {len(qw)} quick-win queries for top-3 rankings</li>')
    lines.append('      <li>Review and improve content depth for underperforming pages</li>')
    lines.append('      <li>Set up ongoing monitoring and reporting cadence</li>')
    lines.append('    </ul>')
    lines.append('  </div>')
    lines.append('')

    lines.append('</div>')
    return "\n".join(lines)


def _build_methodology_footer(domain, timestamp):
    """Build the Data Sources & Methodology footer section."""
    return (
        f'\n<!-- {"=" * 55} DATA SOURCES & METHODOLOGY {"=" * 3} -->\n'
        f'<div class="section" style="text-align: center; padding-top: 15mm;">\n'
        f'  <hr class="divider">\n'
        f'  <h3 style="text-align: left;">Data Sources &amp; Methodology</h3>\n'
        f'  <table>\n'
        f'    <thead>\n'
        f'      <tr><th>Source</th><th>Description</th><th>Update Frequency</th></tr>\n'
        f'    </thead>\n'
        f'    <tbody>\n'
        f'      <tr><td>PageSpeed Insights API</td>\n'
        f'          <td>Lighthouse lab audit (mobile emulation, Moto G Power, slow 4G)</td>\n'
        f'          <td>Real-time</td></tr>\n'
        f'      <tr><td>Chrome UX Report (CrUX)</td>\n'
        f'          <td>28-day rolling field data from real Chrome users</td>\n'
        f'          <td>Daily ~04:00 UTC</td></tr>\n'
        f'      <tr><td>CrUX History API</td>\n'
        f'          <td>25-week p75 trend data per metric</td>\n'
        f'          <td>Weekly</td></tr>\n'
        f'      <tr><td>Google Search Console</td>\n'
        f'          <td>Search Analytics (clicks, impressions, CTR, position)</td>\n'
        f'          <td>2-3 day lag</td></tr>\n'
        f'      <tr><td>URL Inspection API</td>\n'
        f'          <td>Per-URL index status, coverage state, crawl info</td>\n'
        f'          <td>Real-time (2,000/day)</td></tr>\n'
        f'    </tbody>\n'
        f'  </table>\n'
        f'  <p style="color: #94a3b8; font-size: 9pt; margin-top: 5mm;">\n'
        f'    Report generated by SEO AASHAN &mdash; Google SEO Intelligence Skill &mdash; '
        f'{timestamp}<br>\n'
        f'    Methodology based on Google Web Vitals thresholds, Search Console documentation, '
        f'and Lighthouse scoring algorithms.\n'
        f'  </p>\n'
        f'</div>\n'
    )


def _build_crawl_single_report(domain, timestamp, data):
    """Build the Single Page SEO crawl report HTML body."""
    url = data.get("url", f"https://{domain}")
    seo_score = data.get("seoScore", 0)
    findings = data.get("findings", {})

    lines = []

    # 1. Executive Summary
    lines.append('\n<!-- ======================================================= 1. EXECUTIVE SUMMARY === -->')
    lines.append('<div class="section">')
    lines.append('  <div class="section-header">')
    lines.append('    <h2>1. Executive Summary</h2>')
    lines.append(f'    <div class="section-score" style="color: {_score_color(seo_score)};">{seo_score}/100</div>')
    lines.append('  </div>')
    lines.append(f'  <p>Single Page SEO Audit for <strong>{url}</strong> conducted on {timestamp}.</p>')

    # Metric cards row
    word_count = data.get("wordCount", 0)
    img_count = findings.get("imageCount", 0)
    missing_alt = findings.get("missingAltCount", 0)
    internal_links = findings.get("internalLinks", 0)
    external_links = findings.get("externalLinks", 0)
    schema_count = findings.get("schemaCount", 0)

    lines.append('  <div class="four-col">')
    lines.append(f'    <div class="col">{_metric_card(word_count, "Word Count", BRAND["primary"])}</div>')
    lines.append(f'    <div class="col">{_metric_card(img_count, "Total Images", BRAND["secondary"])}</div>')
    lines.append(f'    <div class="col">{_metric_card(f"{internal_links} / {external_links}", "Internal/External Links", BRAND["accent"])}</div>')
    lines.append(f'    <div class="col">{_metric_card(schema_count, "Schema Objects", BRAND["success"])}</div>')
    lines.append('  </div>')

    # Summary of tags check box
    lines.append('  <div class="highlight" style="margin-top: 6mm;">')
    lines.append('    <strong>Page Profile Summary:</strong>')
    lines.append('    <ul>')
    lines.append(f'      <li><strong>Title Tag</strong>: {"Present" if findings.get("hasTitle") else "Missing"}</li>')
    lines.append(f'      <li><strong>Meta Description</strong>: {"Present" if findings.get("hasDesc") else "Missing"}</li>')
    lines.append(f'      <li><strong>H1 Count</strong>: {findings.get("h1Count", 0)} (optimal: 1)</li>')
    lines.append(f'      <li><strong>Images Missing Alt Tags</strong>: {missing_alt}</li>')
    lines.append('    </ul>')
    lines.append('  </div>')
    lines.append('</div>')

    # 2. Gap Analysis
    lines.append('\n<!-- ======================================================= 2. GAP ANALYSIS === -->')
    lines.append('<div class="section">')
    lines.append('  <div class="section-header">')
    lines.append('    <h2>2. On-Page Gap Analysis Details</h2>')
    lines.append('  </div>')
    lines.append('  <table>')
    lines.append('    <thead>')
    lines.append('      <tr><th>SEO Factor</th><th>Status</th><th>Detected Value / Insight</th></tr>')
    lines.append('    </thead>')
    lines.append('    <tbody>')

    # Title row
    title_ok = findings.get("hasTitle")
    title_status = '<span class="status-pass">PASS</span>' if title_ok else '<span class="status-fail">FAIL</span>'
    title_text = findings.get("titleText", "")
    title_val = f'"{title_text}"' if title_text else "Missing &lt;title&gt; element in page &lt;head&gt;."
    lines.append(f'      <tr><td><strong>Title Tag</strong></td><td>{title_status}</td><td>{title_val}</td></tr>')

    # Meta Desc row
    desc_ok = findings.get("hasDesc")
    desc_status = '<span class="status-pass">PASS</span>' if desc_ok else '<span class="status-warn">WARN</span>'
    desc_text = findings.get("descText", "")
    desc_val = f'"{desc_text}"' if desc_text else "Missing &lt;meta name=\'description\'&gt; tag."
    lines.append(f'      <tr><td><strong>Meta Description</strong></td><td>{desc_status}</td><td>{desc_val}</td></tr>')

    # H1 count row
    h1_count = findings.get("h1Count", 0)
    h1_status = '<span class="status-pass">PASS</span>' if h1_count == 1 else '<span class="status-warn">WARN</span>'
    h1_val = f"Found {h1_count} tags. Page must contain exactly one h1 tag for clean hierarchy."
    lines.append(f'      <tr><td><strong>H1 Headings</strong></td><td>{h1_status}</td><td>{h1_val}</td></tr>')

    # Image Alt row
    alt_status = '<span class="status-pass">PASS</span>' if missing_alt == 0 else '<span class="status-warn">WARN</span>'
    alt_val = f"{img_count} images found. {missing_alt} images are missing alternative text attributes."
    lines.append(f'      <tr><td><strong>Image Alt Attributes</strong></td><td>{alt_status}</td><td>{alt_val}</td></tr>')

    # Schema row
    schema_status = '<span class="status-pass">PASS</span>' if schema_count > 0 else '<span class="status-warn">WARN</span>'
    schema_val = f"Detected {schema_count} structured data markup block(s) (JSON-LD or microdata)."
    lines.append(f'      <tr><td><strong>Schema Structured Data</strong></td><td>{schema_status}</td><td>{schema_val}</td></tr>')

    lines.append('    </tbody>')
    lines.append('  </table>')
    lines.append('</div>')

    # 3. Action Plan & Recommendations
    lines.append('\n<!-- ======================================================= 3. RECOMMENDATIONS === -->')
    lines.append('<div class="section">')
    lines.append('  <div class="section-header">')
    lines.append('    <h2>3. Priority Recommendations</h2>')
    lines.append('  </div>')

    rec_count = 0
    if not title_ok:
        rec_count += 1
        lines.append('  <div class="action-item critical">')
        lines.append(f'    <h4>{rec_count}. Inject Target Title Tag <span class="effort">Effort: Low (15 min)</span></h4>')
        lines.append('    <p><strong>Rationale</strong>: The page completely lacks a title tag. Google will auto-generate one from random links, harming click-through rates.</p>')
        lines.append('    <p><strong>Remedy</strong>: Place a descriptive, keyword-rich title between 50-60 characters inside <code>&lt;title&gt;&lt;/title&gt;</code> tags in the head block.</p>')
        lines.append('  </div>')

    if not desc_ok:
        rec_count += 1
        lines.append('  <div class="action-item high">')
        lines.append(f'    <h4>{rec_count}. Add Compelling Meta Description <span class="effort">Effort: Low (20 min)</span></h4>')
        lines.append('    <p><strong>Rationale</strong>: Search snippets will use fallback body fragments which are often not persuasive.</p>')
        lines.append('    <p><strong>Remedy</strong>: Write a meta description (120-155 characters) that includes a call-to-action to maximize click-throughs.</p>')
        lines.append('  </div>')

    if h1_count != 1:
        rec_count += 1
        lines.append('  <div class="action-item medium">')
        lines.append(f'    <h4>{rec_count}. Fix H1 Tag Heading count <span class="effort">Effort: Low (10 min)</span></h4>')
        lines.append(f'    <p><strong>Rationale</strong>: The crawler detected {h1_count} h1 headings. Google expects a single primary theme H1 at the top of the body block.</p>')
        lines.append('    <p><strong>Remedy</strong>: Consolidate the structure so only the main article title uses the h1 tag, converting secondary headings to h2/h3 tags.</p>')
        lines.append('  </div>')

    if missing_alt > 0:
        rec_count += 1
        lines.append('  <div class="action-item medium">')
        lines.append(f'    <h4>{rec_count}. Fix {missing_alt} Missing Image Alt Tags <span class="effort">Effort: Medium (1-2 hrs)</span></h4>')
        lines.append('    <p><strong>Rationale</strong>: Alternative text makes visual elements crawlable for Google Image Search and is required for WCAG accessibility.</p>')
        lines.append('    <p><strong>Remedy</strong>: Add descriptive alt attributes to all indexable images (e.g. <code>alt="Dashboard UI with organic performance charts"</code>).</p>')
        lines.append('  </div>')

    if schema_count == 0:
        rec_count += 1
        lines.append('  <div class="action-item low">')
        lines.append(f'    <h4>{rec_count}. Deploy Schema.org Markup <span class="effort">Effort: Medium (1 hr)</span></h4>')
        lines.append('    <p><strong>Rationale</strong>: Rich results require structured schema markup (like Article, Product, or FAQ) to unlock high-CTR search features.</p>')
        lines.append('    <p><strong>Remedy</strong>: Use the utility console\'s Schema Markup view to generate and embed clean JSON-LD blocks.</p>')
        lines.append('  </div>')

    if rec_count == 0:
        lines.append('  <div class="success-box"><strong>No on-page optimization gaps detected!</strong> This page is properly optimized with meta tags and structural semantic elements.</div>')

    lines.append('</div>')

    return "\n".join(lines)


def _build_crawl_full_report(domain, timestamp, data, chart_paths):
    """Build the Comprehensive Website crawl audit report HTML body."""
    url = data.get("url", f"https://{domain}")
    seo_score = data.get("seoScore", 0)
    crawled_count = data.get("crawledCount", 0)
    sitemap_status = data.get("sitemapStatus", "missing").upper()
    has_llms = "Found" if data.get("hasLlmsTxt") else "Missing"

    lines = []

    # 1. Executive Summary
    lines.append('\n<!-- ======================================================= 1. EXECUTIVE SUMMARY === -->')
    lines.append('<div class="section">')
    lines.append('  <div class="section-header">')
    lines.append('    <h2>1. Executive Summary</h2>')
    lines.append(f'    <div class="section-score" style="color: {_score_color(seo_score)};">{seo_score}/100</div>')
    lines.append('  </div>')
    lines.append(f'  <p>Comprehensive website crawler analysis of <strong>{url}</strong> conducted on {timestamp}. '
                 f'The site was audited using crawl rules checking internal structure, E-E-A-T pages, '
                 f'sitemap layout, and robots guidelines.</p>')

    # Metric cards row
    lines.append('  <div class="four-col">')
    lines.append(f'    <div class="col">{_metric_card(f"{seo_score}/100", "Health Score", _score_color(seo_score))}</div>')
    lines.append(f'    <div class="col">{_metric_card(crawled_count, "Pages Crawled", BRAND["primary"])}</div>')
    lines.append(f'    <div class="col">{_metric_card(sitemap_status, "XML Sitemap Status", BRAND["success"] if "PRESENT" in sitemap_status else BRAND["warning"])}</div>')
    lines.append(f'    <div class="col">{_metric_card(has_llms, "llms.txt Crawlable", BRAND["success"] if "Found" in has_llms else BRAND["muted"])}</div>')
    lines.append('  </div>')

    # E-E-A-T status box
    contacts_found = data.get("contactsFound", False)
    privacy_found = data.get("privacyFound", False)
    eeat_box = "success-box" if (contacts_found and privacy_found) else "highlight"
    lines.append(f'  <div class="{eeat_box}" style="margin-top: 6mm;">')
    lines.append('    <strong>E-E-A-T Accessibility:</strong>')
    lines.append('    <ul>')
    lines.append(f'      <li><strong>Contact / Location page</strong>: {"Present" if contacts_found else "Missing (reduces trust signals)"}</li>')
    lines.append(f'      <li><strong>Privacy / Terms Policy</strong>: {"Present" if privacy_found else "Missing (affects compliance and domain authority)"}</li>')
    lines.append('    </ul>')
    lines.append('  </div>')
    lines.append('</div>')

    # 2. Category Breakdown & Visual Chart
    lines.append('\n<!-- ======================================================= 2. CATEGORY SCORES === -->')
    lines.append('<div class="section">')
    lines.append('  <div class="section-header">')
    lines.append('    <h2>2. Category Weight Contributions</h2>')
    lines.append('  </div>')

    # Insert Matplotlib Categories Chart
    categories_chart_path = chart_paths.get("crawl_categories_path", "")
    if categories_chart_path:
        lines.append(f'  <div class="chart-container">')
        lines.append(f'    <img src="file://{categories_chart_path}" style="width: 80%;" alt="Crawl audit category breakdown">')
        lines.append(f'    <div class="chart-caption">Figure 1: Crawl health breakdown across the 7 critical search factors.</div>')
        lines.append(f'  </div>')

    lines.append('  <table>')
    lines.append('    <thead>')
    lines.append('      <tr><th>Audit Category</th><th>Score</th><th>Weight</th><th>Health Contribution</th></tr>')
    lines.append('    </thead>')
    lines.append('    <tbody>')

    categories_weights = [
        ("technical", "Technical SEO", "22%"),
        ("content", "Content Quality", "23%"),
        ("onpage", "On-Page SEO", "20%"),
        ("schema", "Schema / Structured Data", "10%"),
        ("performance", "Performance (CWV)", "10%"),
        ("geo", "AI Search Readiness (GEO)", "10%"),
        ("images", "Images alt coverage", "5%"),
    ]

    scores = data.get("categoryScores", {})
    for key, label, weight in categories_weights:
        score = scores.get(key, 0)
        contrib = score * float(weight.replace("%", "")) / 100
        cls = "status-pass" if score >= 90 else ("status-warn" if score >= 50 else "status-fail")
        lines.append(f'      <tr><td>{label}</td><td class="{cls}">{score}/100</td><td>{weight}</td><td>{contrib:.2f}</td></tr>')

    lines.append(f'      <tr style="font-weight: bold; border-top: 2px solid {BRAND["grid"]};">')
    lines.append(f'        <td>Total Overall Index</td><td></td><td>100%</td><td style="color: {BRAND["success"]};">{seo_score}/100</td>')
    lines.append('      </tr>')
    lines.append('    </tbody>')
    lines.append('  </table>')
    lines.append('</div>')

    # 3. Prioritized Checklist
    lines.append('\n<!-- ======================================================= 3. FINDINGS CHECKLIST === -->')
    lines.append('<div class="section">')
    lines.append('  <div class="section-header">')
    lines.append('    <h2>3. Prioritized Audit Findings</h2>')
    lines.append('  </div>')
    lines.append('  <p>The crawling rules identified the following optimization gaps, sorted by severity. Address red and orange items first.</p>')

    findings_list = data.get("findings", [])

    criticals = [f for f in findings_list if f.get("severity") == "Critical"]
    highs = [f for f in findings_list if f.get("severity") == "High"]
    mediums = [f for f in findings_list if f.get("severity") == "Medium"]
    lows = [f for f in findings_list if f.get("severity") == "Low"]

    def format_finding_item(f, priority_lbl, item_class):
        return (
            f'  <div class="action-item {item_class}">\n'
            f'    <h4><span class="priority-tag priority-{item_class}">{priority_lbl}</span> {f.get("message", "")}</h4>\n'
            f'    <p><strong>Category</strong>: {f.get("category", "")} | <strong>Rule</strong>: <code>{f.get("rule", "")}</code></p>\n'
            f'    <p><strong>Remedy</strong>: {f.get("remedy", "")}</p>\n'
            f'  </div>'
        )

    item_idx = 0
    if criticals:
        lines.append('  <h3>🔴 Critical Fixes (Fix Immediately)</h3>')
        for f in criticals:
            item_idx += 1
            lines.append(format_finding_item(f, "CRITICAL", "critical"))
        lines.append('')

    if highs:
        lines.append('  <h3>🟠 High Priority Tasks (Fix Within 1 Week)</h3>')
        for f in highs:
            item_idx += 1
            lines.append(format_finding_item(f, "HIGH", "high"))
        lines.append('')

    if mediums:
        lines.append('  <h3>🟡 Medium Priority Items (Fix Within 1 Month)</h3>')
        for f in mediums:
            item_idx += 1
            lines.append(format_finding_item(f, "MEDIUM", "medium"))
        lines.append('')

    if lows:
        lines.append('  <h3>🟢 Low Priority / Long-term Opportunities</h3>')
        for f in lows:
            item_idx += 1
            lines.append(format_finding_item(f, "LOW", "low"))
        lines.append('')

    if item_idx == 0:
        lines.append('  <div class="success-box"><strong>No optimization issues found!</strong> This site is properly tuned according to all Tier 4 crawl policies.</div>')

    lines.append('</div>')

    # 4. Crawled Pages Directory
    lines.append('\n<!-- ======================================================= 4. CRAWLED PAGES DIRECTORY === -->')
    lines.append('<div class="section">')
    lines.append('  <div class="section-header">')
    lines.append('    <h2>4. Crawled Pages Index</h2>')
    lines.append('  </div>')
    lines.append('  <p>Detailed report index of crawl statistics. Shows status code, word count, responsive load time, and page title (capped at 30 pages in preview).</p>')
    lines.append('  <table>')
    lines.append('    <thead>')
    lines.append('      <tr><th>Crawled URL Path</th><th>Status</th><th>Words</th><th>Response Time</th><th>Page Title</th></tr>')
    lines.append('    </thead>')
    lines.append('    <tbody>')

    pages = data.get("crawledPages", [])
    for p in pages[:30]:
        status = p.get("statusCode", 0)
        status_cls = "status-pass" if status == 200 else "status-fail"
        latency = p.get("responseTime", 0)
        lat_cls = "status-pass" if latency <= 0.8 else ("status-warn" if latency <= 1.5 else "status-fail")
        url_display = p.get("url", "").replace(url, "") or "/"

        lines.append(
            f'      <tr>'
            f'<td style="word-break:break-all; font-family: monospace; font-size: 8pt;">{url_display}</td>'
            f'<td class="{status_cls}" style="text-align:center;">{status}</td>'
            f'<td style="text-align:right;">{p.get("wordCount", 0)}</td>'
            f'<td class="{lat_cls}" style="text-align:right;">{latency:.2f}s</td>'
            f'<td style="font-size: 8pt; color: {BRAND["secondary"]};">{p.get("title", "")[:45]}</td>'
            f'</tr>'
        )

    lines.append('    </tbody>')
    lines.append('  </table>')
    if len(pages) > 30:
        lines.append(f'  <p style="font-size: 8.5pt; font-style: italic; color: {BRAND["muted"]}; text-align: center;">... and {len(pages) - 30} additional page(s) crawled and verified.</p>')
    lines.append('</div>')

    # 5. Roadmap Timeline
    lines.append('\n<!-- ======================================================= 5. TIMELINE ROADMAP === -->')
    lines.append('<div class="section">')
    lines.append('  <div class="section-header">')
    lines.append('    <h2>5. Implementation Roadmap</h2>')
    lines.append('  </div>')
    lines.append('  <p>Structured schedule for applying the recommended improvements sequentially.</p>')

    lines.append('  <div class="roadmap-phase">')
    lines.append('    <h4>Phase 1: Week 1 &mdash; Technical Cleanup &amp; Quick Wins</h4>')
    lines.append('    <ul>')
    if "missing-xml-sitemap" in [f.get("rule") for f in findings_list]:
        lines.append('      <li>Create /sitemap.xml and register it in robots.txt</li>')
    if criticals:
        lines.append('      <li>Resolve all Critical severity checklist violations</li>')
    if not contacts_found or not privacy_found:
        lines.append('      <li>Establish Contact and Privacy Policy pages to satisfy E-E-A-T heuristics</li>')
    lines.append('      <li>Improve title and description tags on the homepage</li>')
    lines.append('    </ul>')
    lines.append('  </div>')

    lines.append('  <div class="roadmap-phase">')
    lines.append('    <h4>Phase 2: Weeks 2&ndash;3 &mdash; Page-level Content Depth</h4>')
    lines.append('    <ul>')
    if "thin-content" in [f.get("rule") for f in findings_list]:
        lines.append('      <li>Expand content depth on low-wordcount thin pages</li>')
    if highs:
        lines.append('      <li>Address High severity on-page tag issues</li>')
    if "missing-canonical" in [f.get("rule") for f in findings_list]:
        lines.append('      <li>Deploy canonical headers on missing page directories</li>')
    lines.append('    </ul>')
    lines.append('  </div>')

    lines.append('  <div class="roadmap-phase">')
    lines.append('    <h4>Phase 3: Week 4 &mdash; Schema &amp; Rich Snippets Validation</h4>')
    lines.append('    <ul>')
    lines.append('      <li>Generate schema markup files for all core landing services</li>')
    lines.append('      <li>Validate JSON-LD blocks structure using search crawlers</li>')
    lines.append('      <li>Setup ongoing SEO drift baseline tracker and scheduling rules</li>')
    lines.append('    </ul>')
    lines.append('  </div>')

    lines.append('</div>')

    return "\n".join(lines)


# ─── Report Assemblers ───────────────────────────────────────────────────────

def generate_report(report_type, data, domain, output_dir, output_format="pdf"):
    """
    Generate a complete professional PDF/HTML report.

    Args:
        report_type: 'cwv-audit', 'gsc-performance', 'indexation', or 'full'.
        data: Dictionary with all input data.
        domain: Domain name for the report header.
        output_dir: Directory for output files.
        output_format: 'pdf', 'html', or 'both'.

    Returns:
        Dictionary with output paths.
    """
    output_dir = Path(output_dir)
    charts_dir = output_dir / "charts"
    charts_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%B %d, %Y")
    timestamp_short = datetime.now().strftime("%Y-%m-%d %H:%M")
    result = {"report_type": report_type, "domain": domain, "files": [], "error": None}

    # ── Generate Charts ──────────────────────────────────────────────────────

    chart_paths = {}

    if report_type in ("cwv-audit", "full"):
        psi = data.get("psi", data)
        mobile = psi.get("psi", {}).get("mobile", psi) if isinstance(psi, dict) else {}
        path = chart_lighthouse_gauges(mobile, charts_dir)
        if path:
            chart_paths["gauges_path"] = path

        crux = data.get("crux", {})
        path = chart_cwv_distributions({"crux": crux} if crux else data, charts_dir)
        if path:
            chart_paths["distributions_path"] = path

        history = data.get("crux_history", {})
        if history and not history.get("error"):
            path = chart_cwv_timeline(history, charts_dir)
            if path:
                chart_paths["timeline_path"] = path

    if report_type in ("gsc-performance", "full"):
        gsc = data.get("gsc", data)
        path = chart_top_queries(gsc, charts_dir)
        if path:
            chart_paths["top_queries_path"] = path

    if report_type in ("indexation", "full"):
        inspect = data.get("inspection", data)
        path = chart_index_status(inspect, charts_dir)
        if path:
            chart_paths["index_status_path"] = path

    if report_type == "crawl-full":
        path = chart_crawl_categories(data, charts_dir)
        if path:
            chart_paths["crawl_categories_path"] = path

    # ── Build HTML Sections ──────────────────────────────────────────────────

    sections = []
    fig_num = 1

    # ── CRAWL-SINGLE report ──────────────────────────────────────────────────
    if report_type == "crawl-single":
        seo_score = data.get("seoScore", 0)
        sections.append(_build_title_page(
            domain, "Single Page SEO Audit",
            "On-Page Technical &amp; Content Quality",
            score=seo_score,
            score_label="SEO Health Index",
            meta_items=[timestamp, "Headless Crawl Auditor"],
        ))
        sections.append(_build_crawl_single_report(domain, timestamp, data))
        sections.append(_build_methodology_footer(domain, timestamp))

    # ── CRAWL-FULL report ────────────────────────────────────────────────────
    elif report_type == "crawl-full":
        seo_score = data.get("seoScore", 0)
        crawled_count = data.get("crawledCount", 0)
        sections.append(_build_title_page(
            domain, "Website Crawl Audit Report",
            "Comprehensive Architecture &amp; On-Page Review",
            score=seo_score,
            score_label="SEO Health Index",
            meta_items=[timestamp, f"BFS Crawler ({crawled_count} pages)"],
        ))

        toc_sections = [
            {"num": 1, "title": "Executive Summary", "subs": ["Key Metrics &amp; Trust Accessibility"]},
            {"num": 2, "title": "Category Weight Contributions", "subs": ["Crawl Health Breakdown Chart"]},
            {"num": 3, "title": "Prioritized Audit Findings", "subs": ["Actionable Optimization Gaps"]},
            {"num": 4, "title": "Crawled Pages Index", "subs": ["Status, Wordcount &amp; Performance Index"]},
            {"num": 5, "title": "Implementation Roadmap", "subs": ["Week 1", "Weeks 2-3", "Week 4"]},
            {"num": 6, "title": "Data Sources &amp; Methodology", "subs": []},
        ]
        sections.append(_build_toc(toc_sections))
        sections.append(_build_crawl_full_report(domain, timestamp, data, chart_paths))
        sections.append(_build_methodology_footer(domain, timestamp))

    # ── CWV-AUDIT report ─────────────────────────────────────────────────────
    elif report_type == "cwv-audit":
        mobile = data.get("psi", data)
        if isinstance(mobile, dict):
            mobile = mobile.get("psi", {}).get("mobile", mobile)
        perf_score = mobile.get("lighthouse_scores", {}).get("performance") if isinstance(mobile, dict) else None

        sections.append(_build_title_page(
            domain, "Core Web Vitals Audit",
            "Performance &amp; User Experience Analysis",
            score=perf_score,
            score_label="Lighthouse Performance Score",
            meta_items=[timestamp, "PageSpeed Insights + CrUX"],
        ))

        # TOC
        toc_sections = [
            {"num": 1, "title": "Executive Summary", "subs": [
                "Key Metrics &amp; Critical Issues",
            ]},
            {"num": 2, "title": "Core Web Vitals &amp; Performance", "score": perf_score, "subs": [
                "Lighthouse Scores",
                "Lab Metrics",
                "CrUX Field Data",
                "Failed Audits &amp; SEO Checks",
            ]},
            {"num": 3, "title": "Recommendations", "subs": [
                "Prioritized Action Items",
            ]},
            {"num": 4, "title": "Data Sources &amp; Methodology", "subs": []},
        ]
        sections.append(_build_toc(toc_sections))

        sections.append(_build_executive_summary(domain, timestamp, data, report_type))

        cwv_html, fig_num = _build_cwv_section(
            data, data.get("crux", {}), chart_paths,
            data.get("crux_history"), section_num=2,
        )
        sections.append(cwv_html)

        sections.append(_build_recommendations(data, section_num=3))
        sections.append(_build_methodology_footer(domain, timestamp))

    # ── GSC-PERFORMANCE report ───────────────────────────────────────────────
    elif report_type == "gsc-performance":
        gsc = data.get("gsc", data)
        clicks = gsc.get("totals", {}).get("clicks", 0)

        sections.append(_build_title_page(
            domain, "Search Console Performance",
            "Google Search Analytics Report",
            score=f"{clicks:,}",
            score_label="Total Clicks",
            meta_items=[timestamp, "Google Search Console API"],
        ))

        toc_sections = [
            {"num": 1, "title": "Executive Summary", "subs": [
                "Key Metrics &amp; Quick Wins",
            ]},
            {"num": 2, "title": "Search Console Performance", "subs": [
                "Key Metrics",
                "Top Queries by Clicks",
                "Query Detail Table",
                "Position Analysis &amp; Quick Wins",
            ]},
            {"num": 3, "title": "Recommendations", "subs": [
                "Prioritized Action Items",
            ]},
            {"num": 4, "title": "Data Sources &amp; Methodology", "subs": []},
        ]
        sections.append(_build_toc(toc_sections))

        sections.append(_build_executive_summary(domain, timestamp, data, report_type))

        gsc_html, fig_num = _build_gsc_section(gsc, chart_paths, section_num=2)
        sections.append(gsc_html)

        sections.append(_build_recommendations(data, section_num=3))
        sections.append(_build_methodology_footer(domain, timestamp))

    # ── INDEXATION report ────────────────────────────────────────────────────
    elif report_type == "indexation":
        inspect = data.get("inspection", data)
        total = inspect.get("total", 0)

        sections.append(_build_title_page(
            domain, "Indexation Status Report",
            "URL Index Coverage Analysis",
            score=total,
            score_label="URLs Inspected",
            meta_items=[timestamp, "URL Inspection API"],
        ))

        toc_sections = [
            {"num": 1, "title": "Executive Summary", "subs": [
                "Index Coverage Overview",
            ]},
            {"num": 2, "title": "Indexation Status", "subs": [
                "Index Coverage Overview",
                "Per-URL Results",
            ]},
            {"num": 3, "title": "Recommendations", "subs": [
                "Prioritized Action Items",
            ]},
            {"num": 4, "title": "Data Sources &amp; Methodology", "subs": []},
        ]
        sections.append(_build_toc(toc_sections))

        sections.append(_build_executive_summary(domain, timestamp, data, report_type))

        idx_html, fig_num = _build_indexation_section(inspect, chart_paths, section_num=2)
        sections.append(idx_html)

        sections.append(_build_recommendations(data, section_num=3))
        sections.append(_build_methodology_footer(domain, timestamp))

    # ── FULL report ──────────────────────────────────────────────────────────
    elif report_type == "full":
        psi = data.get("psi", {})
        mobile = psi.get("psi", {}).get("mobile", psi) if isinstance(psi, dict) else {}
        perf_score = mobile.get("lighthouse_scores", {}).get("performance") if isinstance(mobile, dict) else None

        sections.append(_build_title_page(
            domain, "Google SEO Intelligence Report",
            "Comprehensive Analysis",
            score=perf_score,
            score_label="Lighthouse Performance Score" if perf_score else None,
            meta_items=[timestamp, "All Google APIs"],
        ))

        # Build TOC dynamically based on available data
        toc_sections = [
            {"num": 1, "title": "Executive Summary", "subs": [
                "Key Metrics, Critical Issues &amp; Quick Wins",
            ]},
        ]
        sec_num = 2
        if data.get("psi") or data.get("crux"):
            toc_sections.append({
                "num": sec_num, "title": "Core Web Vitals &amp; Performance",
                "score": perf_score, "subs": [
                    "Lighthouse Scores &amp; Lab Metrics",
                    "CrUX Field Data &amp; Trends",
                    "Failed Audits &amp; Opportunities",
                ],
            })
            sec_num += 1
        if data.get("gsc"):
            toc_sections.append({
                "num": sec_num, "title": "Search Console Performance", "subs": [
                    "Key Metrics &amp; Top Queries",
                    "Position Analysis &amp; Quick Wins",
                ],
            })
            sec_num += 1
        if data.get("inspection"):
            toc_sections.append({
                "num": sec_num, "title": "Indexation Status", "subs": [
                    "Index Coverage &amp; Per-URL Results",
                ],
            })
            sec_num += 1
        toc_sections.append({
            "num": sec_num, "title": "Recommendations", "subs": [
                "Prioritized Action Items",
            ],
        })
        rec_num = sec_num
        sec_num += 1
        toc_sections.append({
            "num": sec_num, "title": "Data Sources &amp; Methodology", "subs": [],
        })

        sections.append(_build_toc(toc_sections))
        sections.append(_build_executive_summary(domain, timestamp, data, report_type))

        current_sec = 2
        if data.get("psi") or data.get("crux"):
            cwv_html, fig_num = _build_cwv_section(
                data.get("psi", {}), data.get("crux", {}), chart_paths,
                data.get("crux_history"), section_num=current_sec,
            )
            sections.append(cwv_html)
            current_sec += 1

        if data.get("gsc"):
            gsc_html, fig_num = _build_gsc_section(
                data["gsc"], chart_paths,
                section_num=current_sec, fig_start=fig_num,
            )
            sections.append(gsc_html)
            current_sec += 1

        if data.get("inspection"):
            idx_html, fig_num = _build_indexation_section(
                data["inspection"], chart_paths,
                section_num=current_sec, fig_start=fig_num,
            )
            sections.append(idx_html)
            current_sec += 1

        sections.append(_build_recommendations(data, section_num=rec_num))
        sections.append(_build_methodology_footer(domain, timestamp))

    # ── Assemble Final HTML ──────────────────────────────────────────────────

    css = _build_css(domain)
    body = "\n".join(sections)
    html_content = (
        f'<!DOCTYPE html>\n'
        f'<html lang="en">\n'
        f'<head>\n'
        f'<meta charset="UTF-8">\n'
        f'<style>\n'
        f'{css}\n'
        f'</style>\n'
        f'</head>\n'
        f'<body>\n'
        f'{body}\n'
        f'</body>\n'
        f'</html>\n'
    )

    # ── Write Output Files ───────────────────────────────────────────────────

    safe_domain = domain.replace(":", "_").replace("/", "_")
    base_name = f"Google-SEO-Report-{safe_domain}-{report_type}"

    if output_format in ("html", "both", "all"):
        html_path = output_dir / f"{base_name}.html"
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        result["files"].append(str(html_path))

    if output_format in ("pdf", "both", "all"):
        if not WEASYPRINT_AVAILABLE:
            result["error"] = "PDF generation failed: WeasyPrint dependencies (GTK3/libgobject) missing on Windows. Falling back to HTML."
            # Fallback to write HTML so there is still a generated report file
            html_fallback_path = output_dir / f"{base_name}.html"
            with open(html_fallback_path, "w", encoding="utf-8") as f:
                f.write(html_content)
            result["files"].append(str(html_fallback_path))
        else:
            pdf_path = output_dir / f"{base_name}.pdf"
            try:
                HTML(string=html_content).write_pdf(str(pdf_path))
                result["files"].append(str(pdf_path))
                # Post-generation review
                review = _review_pdf(str(pdf_path), html_content)
                if review:
                    result["review"] = review
            except Exception as e:
                result["error"] = f"PDF generation failed: {e}"

    if output_format in ("xlsx", "all"):
        xlsx_path = generate_xlsx(data, domain, report_type, output_dir)
        if xlsx_path:
            result["files"].append(xlsx_path)

    return result


def _review_pdf(pdf_path: str, html_content: str) -> dict:
    """
    RULE: Always review the PDF before presenting to the user.
    Check for common rendering issues.
    """
    review = {"issues": [], "page_count": None, "file_size_kb": None}

    # File size
    try:
        size = os.path.getsize(pdf_path)
        review["file_size_kb"] = round(size / 1024, 1)
    except OSError:
        pass

    # Page count (if pypdf available)
    try:
        from pypdf import PdfReader
        reader = PdfReader(pdf_path)
        review["page_count"] = len(reader.pages)
    except ImportError:
        pass

    # HTML-level checks
    import re
    # Check for empty chart containers (img with no src)
    empty_imgs = re.findall(r'<img[^>]*src=""[^>]*>', html_content)
    if empty_imgs:
        review["issues"].append(f"{len(empty_imgs)} empty image tag(s) found")

    # Check for very short sections (might appear as mostly whitespace)
    sections = html_content.split('div class="section"')
    for i, sec in enumerate(sections[1:], 1):
        # Strip HTML tags to get text content
        text_only = re.sub(r'<[^>]+>', '', sec[:2000])
        text_only = re.sub(r'\s+', ' ', text_only).strip()
        if len(text_only) < 50:
            review["issues"].append(f"Section {i} has very little text content ({len(text_only)} chars)")

    # Check for duplicate content
    tables = re.findall(r'<table>.*?</table>', html_content, re.DOTALL)
    if len(tables) != len(set(tables)):
        review["issues"].append("Duplicate tables detected")

    if not review["issues"]:
        review["status"] = "PASS"
    else:
        review["status"] = f"WARN ({len(review['issues'])} issues)"

    return review


# ─── XLSX Export ──────────────────────────────────────────────────────────────

def generate_xlsx(data, domain, report_type, output_dir):
    """Generate Excel workbook from audit data.

    Returns path to generated .xlsx file, or None if openpyxl unavailable.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Warning: openpyxl not installed. Skipping xlsx. Install: pip install openpyxl", file=sys.stderr)
        return None

    wb = Workbook()
    output_dir = Path(output_dir)

    # Brand colors for Excel
    navy_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    cream_fill = PatternFill(start_color="FAF9F7", end_color="FAF9F7", fill_type="solid")
    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    amber_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D6D3CC"),
        right=Side(style="thin", color="D6D3CC"),
        top=Side(style="thin", color="D6D3CC"),
        bottom=Side(style="thin", color="D6D3CC"),
    )

    def _style_header(ws, row=1):
        """Apply navy header styling to the first row."""
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def _auto_width(ws):
        """Auto-fit column widths based on content."""
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    def _severity_fill(severity):
        """Return fill color based on severity string."""
        s = str(severity).lower()
        if s in ("critical", "fail", "high"):
            return red_fill
        if s in ("warning", "warn", "medium"):
            return amber_fill
        if s in ("pass", "good", "low"):
            return green_fill
        return cream_fill

    # ── Summary Sheet ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Google SEO Report", "", ""])
    ws.append(["Domain", domain])
    ws.append(["Report Type", report_type])
    ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws.append([])

    if report_type == "crawl-single":
        findings = data.get("findings", {})
        ws.append(["Single Page SEO Crawl Metrics", ""])
        ws.append(["Metric", "Value"])
        _style_header(ws, ws.max_row)
        ws.append(["Word Count", data.get("wordCount", 0)])
        ws.append(["SEO Health Index", f"{data.get('seoScore', 0)}/100"])
        ws.append(["Total Images", findings.get("imageCount", 0)])
        ws.append(["Images Missing Alt", findings.get("missingAltCount", 0)])
        ws.append(["Internal Links", findings.get("internalLinks", 0)])
        ws.append(["External Links", findings.get("externalLinks", 0)])
        ws.append(["Schema Objects", findings.get("schemaCount", 0)])
        ws.append([])
        
        # On-Page elements sheet
        ws_tags = wb.create_sheet("On-Page Tags")
        ws_tags.append(["On-Page Factor", "Status", "Value / Message"])
        _style_header(ws_tags)
        
        ws_tags.append(["Title Tag", "PASS" if findings.get("hasTitle") else "FAIL", findings.get("titleText", "Missing title tag")])
        ws_tags.append(["Meta Description", "PASS" if findings.get("hasDesc") else "WARN", findings.get("descText", "Missing meta description")])
        
        h1_c = findings.get("h1Count", 0)
        ws_tags.append(["H1 Heading Count", "PASS" if h1_c == 1 else "WARN", f"Found {h1_c} h1 heading tag(s)"])
        
        missing_a = findings.get("missingAltCount", 0)
        ws_tags.append(["Image Alt Tags", "PASS" if missing_a == 0 else "WARN", f"{missing_a} images are missing alternative alt attributes"])
        
        schema_c = findings.get("schemaCount", 0)
        ws_tags.append(["Schema Structured Data", "PASS" if schema_c > 0 else "WARN", f"Found {schema_c} schema definitions"])
        
        for r in range(2, 7):
            cell_status = ws_tags.cell(row=r, column=2)
            cell_status.fill = _severity_fill(cell_status.value)
            cell_status.alignment = Alignment(horizontal="center")
            
        _auto_width(ws_tags)
        
    elif report_type == "crawl-full":
        # Crawl Statistics Summary
        ws.append(["Website Crawl Statistics", ""])
        ws.append(["Stat", "Value"])
        _style_header(ws, ws.max_row)
        ws.append(["Health Index Score", f"{data.get('seoScore', 0)}/100"])
        ws.append(["Pages Crawled Count", data.get("crawledCount", 0)])
        ws.append(["XML Sitemap", data.get("sitemapStatus", "missing").upper()])
        ws.append(["llms.txt configuration", "Found" if data.get("hasLlmsTxt") else "Missing"])
        ws.append(["Contact info present", "Yes" if data.get("contactsFound") else "No"])
        ws.append(["Privacy page present", "Yes" if data.get("privacyFound") else "No"])
        ws.append([])
        
        # Category scores breakdown
        ws.append(["Category Breakdown Scores", ""])
        ws.append(["Category", "Score", "Weight"])
        _style_header(ws, ws.max_row)
        
        categories_list = [
            ("technical", "Technical SEO", "22%"),
            ("content", "Content Quality", "23%"),
            ("onpage", "On-Page SEO", "20%"),
            ("schema", "Schema / Structured Data", "10%"),
            ("performance", "Performance (CWV)", "10%"),
            ("geo", "AI Search Readiness (GEO)", "10%"),
            ("images", "Images alt coverage", "5%"),
        ]
        scores_dict = data.get("categoryScores", {})
        for key, name, weight in categories_list:
            sc = scores_dict.get(key, 0)
            ws.append([name, f"{sc}/100", weight])
            row_idx = ws.max_row
            ws.cell(row=row_idx, column=2).fill = _severity_fill("pass" if sc >= 90 else "warning" if sc >= 50 else "fail")
        ws.append([])
        
        # Findings sheet
        findings_list = data.get("findings", [])
        if findings_list:
            ws_find = wb.create_sheet("Findings")
            ws_find.append(["Category", "Severity", "Rule ID", "Message", "Remedy / Fix"])
            _style_header(ws_find)
            
            sev_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
            sorted_findings = sorted(findings_list, key=lambda f: sev_order.get(f.get("severity", "Low"), 4))
            for f in sorted_findings:
                ws_find.append([
                    f.get("category", ""),
                    f.get("severity", ""),
                    f.get("rule", ""),
                    f.get("message", ""),
                    f.get("remedy", ""),
                ])
                row_idx = ws_find.max_row
                ws_find.cell(row=row_idx, column=2).fill = _severity_fill(f.get("severity", ""))
                ws_find.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
            
            ws_find.auto_filter.ref = f"A1:E{ws_find.max_row}"
            ws_find.freeze_panes = "A2"
            _auto_width(ws_find)
            
        # Crawled Pages sheet
        pages_list = data.get("crawledPages", [])
        if pages_list:
            ws_pages = wb.create_sheet("Crawled Pages")
            ws_pages.append(["URL", "Status Code", "Word Count", "Response Latency", "Page Title"])
            _style_header(ws_pages)
            for p in pages_list:
                ws_pages.append([
                    p.get("url", ""),
                    p.get("statusCode", 0),
                    p.get("wordCount", 0),
                    p.get("responseTime", 0),
                    p.get("title", ""),
                ])
                row_idx = ws_pages.max_row
                status = p.get("statusCode", 0)
                ws_pages.cell(row=row_idx, column=2).fill = _severity_fill("pass" if status == 200 else "fail")
                ws_pages.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
                latency = p.get("responseTime", 0)
                ws_pages.cell(row=row_idx, column=4).fill = _severity_fill("pass" if latency <= 0.8 else "warning" if latency <= 1.5 else "fail")
                
            ws_pages.auto_filter.ref = f"A1:E{ws_pages.max_row}"
            ws_pages.freeze_panes = "A2"
            _auto_width(ws_pages)

    # Add scores if available
    if report_type in ("cwv-audit", "full"):
        psi = data.get("psi", data)
        mobile = psi.get("psi", {}).get("mobile", psi) if isinstance(psi, dict) else {}
        scores = mobile.get("lighthouse_scores", {}) if isinstance(mobile, dict) else {}
        if scores:
            ws.append(["Lighthouse Scores", ""])
            ws.append(["Category", "Score"])
            _style_header(ws, ws.max_row)
            for cat in ["performance", "accessibility", "best_practices", "seo"]:
                val = scores.get(cat) if scores.get(cat) is not None else scores.get(cat.replace("_", "-"))
                if val is not None:
                    row_num = ws.max_row + 1
                    ws.append([cat.replace("_", " ").title(), int(val * 100) if val <= 1 else val])
                    ws.cell(row=row_num, column=2).fill = _severity_fill(
                        "pass" if (val * 100 if val <= 1 else val) >= 90
                        else "warning" if (val * 100 if val <= 1 else val) >= 50
                        else "fail"
                    )
            ws.append([])

    # CWV metrics
    if report_type in ("cwv-audit", "full"):
        crux = data.get("crux", {})
        metrics = crux.get("metrics", {}) if isinstance(crux, dict) else {}
        if metrics:
            ws.append(["Core Web Vitals (Field Data)", ""])
            ws.append(["Metric", "Value", "Rating"])
            _style_header(ws, ws.max_row)
            for metric_name, metric_data in metrics.items():
                if isinstance(metric_data, dict):
                    p75 = metric_data.get("percentile_p75", metric_data.get("p75", ""))
                    rating = metric_data.get("category", "")
                    ws.append([metric_name, p75, rating])
            ws.append([])

    _auto_width(ws)

    # ── GSC Queries Sheet ─────────────────────────────────────────────────────
    gsc = data.get("gsc", {})
    queries = gsc.get("queries", gsc.get("rows", []))
    if queries and isinstance(queries, list):
        ws2 = wb.create_sheet("Queries")
        ws2.append(["Query", "Clicks", "Impressions", "CTR", "Position"])
        _style_header(ws2)
        for row_data in queries[:500]:
            if isinstance(row_data, dict):
                keys = row_data.get("keys", [])
                query = keys[0] if keys else row_data.get("query", "")
                ws2.append([
                    query,
                    row_data.get("clicks", 0),
                    row_data.get("impressions", 0),
                    f"{row_data.get('ctr', 0):.2%}" if isinstance(row_data.get("ctr"), (int, float)) else str(row_data.get("ctr", "")),
                    round(row_data.get("position", 0), 1) if isinstance(row_data.get("position"), (int, float)) else row_data.get("position", ""),
                ])
        ws2.auto_filter.ref = f"A1:E{ws2.max_row}"
        ws2.freeze_panes = "A2"
        _auto_width(ws2)

    # ── GSC Pages Sheet ───────────────────────────────────────────────────────
    pages = gsc.get("pages", [])
    if pages and isinstance(pages, list):
        ws3 = wb.create_sheet("Pages")
        ws3.append(["Page", "Clicks", "Impressions", "CTR", "Position"])
        _style_header(ws3)
        for row_data in pages[:500]:
            if isinstance(row_data, dict):
                keys = row_data.get("keys", [])
                page = keys[0] if keys else row_data.get("page", "")
                ws3.append([
                    page,
                    row_data.get("clicks", 0),
                    row_data.get("impressions", 0),
                    f"{row_data.get('ctr', 0):.2%}" if isinstance(row_data.get("ctr"), (int, float)) else str(row_data.get("ctr", "")),
                    round(row_data.get("position", 0), 1) if isinstance(row_data.get("position"), (int, float)) else row_data.get("position", ""),
                ])
        ws3.auto_filter.ref = f"A1:E{ws3.max_row}"
        ws3.freeze_panes = "A2"
        _auto_width(ws3)

    # ── Indexation Sheet ──────────────────────────────────────────────────────
    inspection = data.get("inspection", {})
    results = inspection.get("results", [])
    if results and isinstance(results, list):
        ws4 = wb.create_sheet("Indexation")
        ws4.append(["URL", "Verdict", "Coverage State", "Indexing State", "Crawled As", "Last Crawl"])
        _style_header(ws4)
        for item in results[:500]:
            if isinstance(item, dict):
                result_data = item.get("inspectionResult", item)
                idx = result_data.get("indexStatusResult", {})
                ws4.append([
                    item.get("url", result_data.get("inspectedUrl", "")),
                    idx.get("verdict", ""),
                    idx.get("coverageState", ""),
                    idx.get("indexingState", ""),
                    idx.get("crawledAs", ""),
                    idx.get("lastCrawlTime", ""),
                ])
        ws4.auto_filter.ref = f"A1:F{ws4.max_row}"
        ws4.freeze_panes = "A2"
        _auto_width(ws4)

    # ── Save ──────────────────────────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d-%H%M")
    filename = f"Google-SEO-Report-{domain}-{timestamp}.xlsx"
    filepath = output_dir / filename
    wb.save(filepath)
    return str(filepath)


# ─── CLI ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Google SEO Report Generator - Professional PDF/HTML reports"
    )
    parser.add_argument(
        "--type", "-t",
        choices=["cwv-audit", "gsc-performance", "indexation", "full", "crawl-single", "crawl-full"],
        required=True,
        help="Report type",
    )
    parser.add_argument("--data", "-d", help="Path to JSON data file (or pipe via stdin)")
    parser.add_argument("--domain", required=True, help="Domain name for the report header")
    parser.add_argument("--output-dir", "-o", default=".", help="Output directory (default: current)")
    parser.add_argument(
        "--format", "-f",
        choices=["pdf", "html", "xlsx", "both", "all"],
        default="pdf",
        help="Output format: pdf, html, xlsx, both (pdf+html), all (pdf+html+xlsx)",
    )
    parser.add_argument("--json", "-j", action="store_true", help="Output metadata as JSON")

    args = parser.parse_args()

    # Load data
    if args.data:
        try:
            with open(args.data, "r") as f:
                data = json.load(f)
        except (json.JSONDecodeError, IOError) as e:
            print(f"Error reading data file: {e}", file=sys.stderr)
            sys.exit(1)
    elif not sys.stdin.isatty():
        try:
            data = json.load(sys.stdin)
        except json.JSONDecodeError as e:
            print(f"Error parsing stdin JSON: {e}", file=sys.stderr)
            sys.exit(1)
    else:
        print("Error: Provide --data file or pipe JSON via stdin.", file=sys.stderr)
        sys.exit(1)

    result = generate_report(
        report_type=args.type,
        data=data,
        domain=args.domain,
        output_dir=args.output_dir,
        output_format=args.format,
    )

    if result.get("error"):
        print(f"Error: {result['error']}", file=sys.stderr)

    if args.json:
        print(json.dumps(result, indent=2))
    else:
        for f in result.get("files", []):
            print(f"Generated: {f}")


if __name__ == "__main__":
    main()
